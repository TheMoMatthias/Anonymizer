"""Second-wave hardening: XXE-safe XML parsing and the short-deny-term backstop."""

from lxml import etree

from anonymizer import xmlsafe
from anonymizer.core import detect_unit
from anonymizer.models import TextUnit
from anonymizer.pipeline import _literal_residual, _with_propagation


def test_xmlsafe_blocks_entity_expansion():
    """Untrusted document XML must not expand entities (billion-laughs DoS / local
    file inclusion). Either the entity is left unresolved or the parse is rejected
    -- never expanded."""
    bomb = (
        '<?xml version="1.0"?><!DOCTYPE lolz [<!ENTITY a "AAAAAAAA">]>'
        "<root>&a;</root>"
    ).encode("utf-8")
    try:
        tree = xmlsafe.fromstring(bomb)
    except etree.XMLSyntaxError:
        return  # rejected outright -> safe
    assert "AAAAAAAA" not in "".join(tree.itertext()), "entity was expanded"


def test_literal_residual_verifies_short_deny_terms(tmp_path):
    """A <4-char value is normally skipped by the backstop (avoids false hits on
    common substrings), but a user-asserted deny term must be verified regardless
    of length."""
    out = tmp_path / "out.txt"
    out.write_text("this still contains ng somewhere", encoding="utf-8")

    assert _literal_residual(out, ["ng"]) == []  # skipped: too short, not a deny term
    assert _literal_residual(out, ["ng"], always_check=["ng"]) == ["ng"]  # deny term -> checked


def test_literal_residual_ignores_values_inside_replacement_tokens(tmp_path):
    """A removed value that survives ONLY as a substring of the tool's OWN
    replacement tokens is not a leak -- it is the anonymized output. Regression
    (spurious HARD-FAIL): an NER-misflagged header word 'Kundennr' (removed as a
    LOCATION) is a substring of the [KUNDENNR_n] tokens that replaced the customer
    NUMBERS, so an unmasked substring scan reported a phantom leak and the fail-loud
    gate refused to write ANY output file."""
    out = tmp_path / "out.txt"
    out.write_text("[KUNDENNR_1] [KUNDENNR_2] [LOCATION_3] anonymisiert", encoding="utf-8")
    assert _literal_residual(out, ["Kundennr"]) == []


def test_literal_residual_still_catches_leak_outside_a_token(tmp_path):
    """The token mask must NOT hide a genuine leak: a removed value present in the
    body (not only inside a replacement token) is still reported."""
    out = tmp_path / "out.txt"
    out.write_text("Herr Mueller [PERSON_1] traf Mueller erneut", encoding="utf-8")
    assert _literal_residual(out, ["Mueller"]) == ["Mueller"]


def test_literal_residual_no_phantom_across_spreadsheet_cells(tmp_path):
    """Regression (spurious HARD-FAIL): _output_text_blob concatenated adjacent
    cells' <v> text with no separator, so two unrelated cells -- or the shared-
    string INDICES that string cells store in <v> -- glued into a phantom digit-run
    that matched a removed customer number and refused to write ANY output. Each
    independent cell must be delimited; text WITHIN a cell must still be contiguous."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1000
    ws["B1"] = 20  # "1000" + "20" would glue into the phantom "100020" across cells
    ws["A2"] = "Kundengeheim"  # a real value living contiguously inside one cell
    path = tmp_path / "phantom.xlsx"
    wb.save(path)

    assert _literal_residual(path, ["100020"]) == []  # phantom across A1|B1 -> not a leak
    assert _literal_residual(path, ["Kundengeheim"]) == ["Kundengeheim"]  # real -> still caught


def test_literal_residual_mask_cannot_swallow_a_bracketed_value(tmp_path):
    """LEAK (B2): the token mask erases every `[A-Z0-9_]`-in-brackets run before
    searching. A removed value that itself looks like one -- a bank case reference
    "[FALL_00219384]", or any deny term with brackets -- was erased from the
    haystack too, so a real surviving leak of exactly that string was invisible to
    the backstop and the file shipped marked verified."""
    out = tmp_path / "out.txt"
    out.write_text("Vorgang [FALL_00219384] weiterhin offen. Bearbeiter [PERSON_1].", encoding="utf-8")

    assert _literal_residual(out, ["[FALL_00219384]"]) == ["[FALL_00219384]"]
    assert _literal_residual(out, ["[FALL_00219384]"], always_check=["[FALL_00219384]"]) == ["[FALL_00219384]"]
    # A bracket-bearing value that is only PART of the surrounding text must be
    # caught too -- the mask must not eat the bracketed head of it.
    out.write_text("Akte [FALL_00219384] Mueller liegt vor.", encoding="utf-8")
    assert _literal_residual(out, ["[FALL_00219384] Mueller"]) == ["[FALL_00219384] Mueller"]
    # ...while the phantom-match suppression the mask exists for still holds.
    assert _literal_residual(out, ["Fall"]) == []


def test_literal_residual_catches_bracket_free_value_inside_a_bracket_run(tmp_path):
    """LEAK (B2): the far more common shape -- a regex/deny recognizer captures the
    bare reference `FALL_00219384`, NOT the surrounding delimiters, so the removed
    value carries no brackets at all. The token mask still erased the whole
    `[FALL_00219384]` run from the haystack, so a file that verbatim contains the
    removed reference was reported clean and shipped."""
    out = tmp_path / "out.txt"
    out.write_text("Vorgang [FALL_00219384] weiterhin offen.", encoding="utf-8")
    assert _literal_residual(out, ["FALL_00219384"]) == ["FALL_00219384"]

    out.write_text("Akte [KND_889912] uebergeben.", encoding="utf-8")
    assert _literal_residual(out, ["KND_889912"], always_check=["KND_889912"]) == ["KND_889912"]


def test_literal_residual_bracketed_value_does_not_resurrect_the_token_phantom(tmp_path):
    """REGRESSION (B2-fp): round 1 protected a bracket-bearing searched value by
    leaving the WHOLE overlapping token run unmasked -- which handed every OTHER
    searched value a free match inside that run, resurrecting the exact
    `Kundennr` inside `[KUNDENNR_1]` false positive the mask exists to suppress
    (a spurious HARD FAIL: no file written at all). Protection must be per-value."""
    out = tmp_path / "out.txt"
    out.write_text("A[KUNDENNR_1]B", encoding="utf-8")

    assert _literal_residual(out, ["Kundennr"]) == []
    assert _literal_residual(out, ["Kundennr", "[KUNDENNR_1]"]) == ["[KUNDENNR_1]"]


# --- B4: document-wide propagation seeding -----------------------------------


class _FakeResult:
    """Minimal stand-in for a presidio RecognizerResult."""

    def __init__(self, entity_type: str, start: int, end: int, score: float):
        self.entity_type = entity_type
        self.start = start
        self.end = end
        self.score = score


class _FakeAnalyzer:
    """Returns fixed spans per exact unit text. Propagation SEEDING is what is
    under test here, so the seed must not depend on spaCy's (model-version
    dependent) opinion about a given string."""

    def __init__(self, spans: dict[str, list[tuple[str, int, int, float]]]):
        self._spans = spans

    def analyze(self, text, language, entities=None, allow_list=None, **kwargs):
        return [_FakeResult(*s) for s in self._spans.get(text, [])]


def test_propagation_seed_is_split_on_newlines(base_config):
    """LEAK (B4i): a seed value spanning a newline -- a multi-line cell, or a
    German address block "Hans Mueller\\nHauptstrasse 5" -- was used verbatim as a
    literal propagation pattern. It matched nothing elsewhere, and because the
    surname was taken as the last whitespace-token of the WHOLE value it became
    "Hauptstrasse", so a bare "Mueller" in another cell was never propagated."""
    text = "Hans Mueller\nHauptstrasse 5"
    units = [TextUnit(id="c1", text=text)]
    analyzer = _FakeAnalyzer({text: [("PERSON", 0, len(text), 0.9)]})

    cfg = _with_propagation({**base_config, "languages": ["de"]}, units, analyzer)
    seeds = {value for _entity, value in cfg.get("propagate", [])}

    assert "Hans Mueller" in seeds, f"newline corrupted the seed: {seeds}"
    assert "Mueller" in seeds, f"surname seed lost to the newline -> bare 'Mueller' leaks: {seeds}"
    assert not any("\n" in s for s in seeds), f"a seed still contains a newline: {seeds}"


def test_shared_surname_is_still_propagated_for_both_people(base_config):
    """DELIBERATE trade-off (B4ii). "Hans Mueller" and "Petra Mueller" both seed
    the bare surname "Mueller", so every bare-surname occurrence -- whichever
    person it belongs to -- maps to the SAME pseudonym, merging two data subjects
    into one identity. That is a real correctness defect, but the alternative
    (suppressing an ambiguous surname seed) LEAKS a real surname into a file the
    tool calls verified. Fail-loud wins: we keep propagating. This test pins that
    choice so nobody "fixes" the merge by re-introducing the leak."""
    texts = ["Hans Mueller hat unterschrieben.", "Petra Mueller hat unterschrieben."]
    units = [TextUnit(id=f"p{i}", text=t) for i, t in enumerate(texts)]
    analyzer = _FakeAnalyzer({t: [("PERSON", 0, t.index(" hat"), 0.9)] for t in texts})

    cfg = _with_propagation({**base_config, "languages": ["de"]}, units, analyzer)
    seeds = {value for _entity, value in cfg.get("propagate", [])}
    assert {"Hans Mueller", "Petra Mueller", "Mueller"} <= seeds, seeds

    # The bare surname in an unrelated cell must still be claimed -> no leak.
    bare = TextUnit(id="cell", text="Mueller")
    findings = detect_unit(_FakeAnalyzer({}), bare, cfg)
    assert [f.value for f in findings] == ["Mueller"]
