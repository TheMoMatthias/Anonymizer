"""Precision regression tests added 2026-07-23 after a false-positive report on
a business/audit-log spreadsheet: ordinary German words ("aber", "abdeckung")
flagged as PII, and findings with leading punctuation fused onto the next word
("-Erstellung", ".iboflow-Dateien"). See docs/run_detection-precision_2026-07-23.md
for the investigation and the agreed fix set.
"""

import openpyxl

from anonymizer.core import (
    _is_digit_bearing_code,
    _is_noise_entity,
    _is_pos_implausible,
    _is_single_lowercase_word,
    _is_structural_nonname,
    detect_unit,
    neutralize_structural_noise,
)
from anonymizer.formats.xlsx_handler import _looks_like_name
from anonymizer.models import TextUnit
from anonymizer.pipeline import scan_document


def test_lowercase_single_word_filter():
    assert _is_single_lowercase_word("aber")
    assert _is_single_lowercase_word("abdeckung")
    assert not _is_single_lowercase_word("Müller")  # capitalized -- a real name shape
    assert not _is_single_lowercase_word("aber schon")  # multi-word: not this filter's job
    assert not _is_single_lowercase_word("")


def test_noise_entity_catches_the_reported_words(analyzer, base_config):
    """Exact regression for the reported false positives, independent of
    whether a crafted test sentence happens to reproduce the model's misfire
    (spaCy's NER is context-sensitive; this pins the filter's own behavior)."""
    lang = "de"
    for word in ("aber", "abdeckung", "und", "Aber"):  # "Aber": capitalized at sentence-start
        assert _is_noise_entity("NER_MISC", word, analyzer, lang), f"{word!r} should be filtered as noise"
    # PERSON is deliberately excluded -- a lowercase surname must stay reachable
    # via other means (honorific/labelled patterns), not silently gated here.
    assert not _is_noise_entity("PERSON", "aber", analyzer, lang)
    # A real org name must not be caught by this filter.
    assert not _is_noise_entity("ORGANIZATION", "Deutsche Bank", analyzer, lang)


def test_common_german_prose_words_not_flagged(analyzer, base_config):
    """The reported false-positive class: ordinary business prose, not a table
    of names."""
    cfg = {**base_config, "languages": ["de"]}
    text = (
        "Die Erstellung der Schnittstellen erfolgt zeitnah, aber die Abdeckung "
        "der Altprozesse ist noch offen und wird im Rahmen der Migration geprüft."
    )
    findings = detect_unit(analyzer, TextUnit("u1", text), cfg)
    flagged = {f.value.lower() for f in findings}
    assert "aber" not in flagged
    assert not any("abdeckung" in v for v in flagged)


def test_leading_punctuation_trimmed_from_fused_token(analyzer, base_config):
    """Real, reproduced case: a short fragment (as a spreadsheet cell or list
    item often is) with no space after a structural character gets its whole
    fused token -- punctuation included -- tagged as an entity by spaCy. The
    leading punctuation must never survive into the finding."""
    cfg = {**base_config, "languages": ["de"]}
    text = "manuelle Migration der .iboflow-Dateien ist zeitaufwendig."
    findings = detect_unit(analyzer, TextUnit("u1", text), cfg)
    assert findings, "expected the model to still tag something here"
    for f in findings:
        assert not f.value[:1] or f.value[0].isalnum(), f"leading punctuation leaked: {f.value!r}"
        assert not f.value.startswith("."), f"leading period leaked: {f.value!r}"


def test_neutralize_structural_noise_is_same_length():
    text = "Migration abgeschlossen.\n-Erstellung neuer Schnittstellen\n-Rollout"
    cleaned = neutralize_structural_noise(text)
    assert len(cleaned) == len(text)
    assert "-Erstellung" not in cleaned
    assert "-Rollout" not in cleaned


def test_neutralize_structural_noise_leaves_real_content_alone():
    # An already-spaced bullet, and a hyphenated word mid-sentence, must survive untouched.
    text = "- Erstellung ist fertig\nEin Müller-Schmidt Termin steht an."
    assert neutralize_structural_noise(text) == text


def test_looks_like_name():
    assert _looks_like_name("Klaus Müller")
    assert _looks_like_name("Klaus von Bergen")
    assert _looks_like_name("Dr. Klaus Müller")
    assert not _looks_like_name("Ein konkreter Ist-Prozess sowie belastbare Pain Points wurden dokumentiert.")
    assert not _looks_like_name("eine lange Liste von unwichtigen Dingen ohne einen echten Namen hier")


def test_name_shape_gate_excludes_prose_paragraph(tmp_path, analyzer, base_config):
    """A column header matching the people-column terms is evidence the COLUMN
    is about people -- not that every cell (a changelog note, a paragraph) is
    a bare name."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Verantwortlich"  # matches "verantwortlich" in _NAME_HEADER_TERMS
    ws["A2"] = (
        "Ein konkreter Ist-Prozess sowie belastbare Pain Points wurden im Rahmen "
        "der Analyse dokumentiert und mit dem Fachbereich abgestimmt."
    )
    path = tmp_path / "changelog.xlsx"
    wb.save(path)

    found = {g.value for g in scan_document(path, analyzer, base_config).all_actionable()}
    assert not any(v.startswith("Ein konkreter") for v in found), (
        f"whole-cell override claimed a prose paragraph as a name: {found}"
    )


def test_name_shape_gate_still_claims_real_name(tmp_path, analyzer, base_config):
    """The gate must not over-restrict: a genuinely name-shaped cell under a
    matching header is still claimed, same as before this change."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Verantwortlich"
    ws["A2"] = "Klaus von Bergen"
    path = tmp_path / "responsible.xlsx"
    wb.save(path)

    found = {g.value for g in scan_document(path, analyzer, base_config).all_actionable()}
    assert "Klaus von Bergen" in found


def test_columns_panel_surfaces_name_override(tmp_path, analyzer, base_config):
    """A column whose header matched the people-column list is flagged as such
    in ColumnInfo, so a coincidental match is visible before Save rather than
    only inferable from an unexpectedly high pii_count."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Verantwortlich"
    ws["A2"] = "Klaus von Bergen"
    ws["B1"] = "Status"
    ws["B2"] = "aktiv"
    path = tmp_path / "columns.xlsx"
    wb.save(path)

    result = scan_document(path, analyzer, base_config)
    by_col = {c.column: c for c in result.columns}
    assert by_col["A"].name_override, "Verantwortlich should trigger the name override flag"
    assert not by_col["B"].name_override, "Status must not be flagged as a name override"


def test_real_pii_still_caught_alongside_ordinary_words(analyzer, base_config):
    """Precision fixes must not cost recall on genuine PII in the same text."""
    cfg = {**base_config, "languages": ["de"]}
    text = "Sehr geehrter Herr Müller, die Abdeckung wurde geprüft, aber noch nicht abgeschlossen."
    findings = detect_unit(analyzer, TextUnit("u1", text), cfg)
    assert any(f.entity_type == "PERSON" and f.value == "Müller" for f in findings)
    flagged = {f.value.lower() for f in findings}
    assert "aber" not in flagged


# --- Second wave: 2026-07-23 follow-up report ("Abgelehnt", "Alle Zielwerte",
# "BP-002"-style codes, and column headers themselves being flagged) ---------


def test_pos_implausible_filters_verb_tagged_person_guess(analyzer, base_config):
    """'Abgelehnt' ('Rejected', a past participle) was tagged PERSON at
    spaCy's flat NER score. PERSON is excluded from the lowercase/stopword
    filter (a real lowercase surname must stay reachable), but the tagger's
    OWN part-of-speech call (VERB, not NOUN/PROPN) disagrees with its NER
    call -- a safe, case-independent signal."""
    cfg = {**base_config, "languages": ["de"]}
    for text in ("Abgelehnt", "Status: Genehmigt", "Der Antrag wurde Abgelehnt."):
        findings = detect_unit(analyzer, TextUnit("u1", text), cfg)
        assert not any(f.value in ("Abgelehnt", "Genehmigt") for f in findings), (text, findings)


def test_pos_implausible_does_not_reject_real_names():
    """Regression guard for the mechanism itself: _is_pos_implausible must
    only fire on entity types in _NER_ENTITIES, and only when NO token in the
    span is noun-class."""
    assert not _is_pos_implausible("IBAN_CODE", 0, 5, None)  # not a free-text NER entity
    assert not _is_pos_implausible("PERSON", 0, 5, None)  # no artifacts -> never filtered


def test_digit_bearing_code_is_never_a_name():
    assert _is_digit_bearing_code("NER_MISC", "BP-002")
    assert _is_digit_bearing_code("PERSON", "Kunde123")
    assert not _is_digit_bearing_code("PERSON", "Müller")
    assert not _is_digit_bearing_code("IBAN_CODE", "DE89370400440532013000")  # not a free-text entity


def test_digit_bearing_project_codes_not_flagged(analyzer, base_config):
    """Real-world shape from the reported spreadsheet: a Project_ID column
    full of 'BP-26-015'-style codes must not be flagged as MISC/PERSON/ORG."""
    cfg = {**base_config, "languages": ["de"]}
    findings = detect_unit(analyzer, TextUnit("u1", "Vorgang BP-26-015 wurde bearbeitet."), cfg)
    assert not any("BP-26-015" in f.value for f in findings)


def test_header_row_not_scanned_as_data(tmp_path, analyzer, base_config):
    """Regression: row 1 (the column's own schema label) used to be scanned
    like any other cell -- a CamelCase/underscore header ("NewValue",
    "Project_ID") reads as a proper noun to NER (tagged PROPN, same as a real
    name), so the header itself showed up as a finding. Row 1 must supply
    header CONTEXT (via _column_headers) but never be its own scannable unit."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Project_ID", "OldValue", "NewValue"])
    ws.append(["BP-26-015", "Ein alter Wert", "Ein neuer Wert"])
    path = tmp_path / "headers.xlsx"
    wb.save(path)

    found = {g.value for g in scan_document(path, analyzer, base_config).all_actionable()}
    assert not found & {"Project_ID", "OldValue", "NewValue"}, f"header label(s) leaked as findings: {found}"


# --- Fourth wave: German-model residual noise (2026-07-23), characterized from
# the real document's export: acronyms, snake_case field IDs, single-letter
# fragments dominated the leftover false positives. ---


def test_structural_nonname_filters_the_measured_noise_shapes():
    # Single letters / 2-char fragments, snake_case identifiers, short acronyms.
    for et in ("PERSON", "NER_MISC", "ORGANIZATION", "LOCATION"):
        assert _is_structural_nonname(et, "S")
        assert _is_structural_nonname(et, "ch")
        assert _is_structural_nonname(et, "Feld_Name")
        assert _is_structural_nonname(et, "Persona_Liste")
        assert _is_structural_nonname(et, "CAPEX")
        assert _is_structural_nonname(et, "RAG")
    # Must NOT filter genuine names/orgs.
    assert not _is_structural_nonname("PERSON", "Müller")
    assert not _is_structural_nonname("PERSON", "Klaus von Bergen")
    assert not _is_structural_nonname("ORGANIZATION", "Deutsche Bank")
    assert not _is_structural_nonname("ORGANIZATION", "FactSet")
    assert not _is_structural_nonname("PERSON", "Yılmaz")
    # Not applied to structured/validated entity types.
    assert not _is_structural_nonname("IBAN_CODE", "DE")


def test_acronyms_and_field_ids_not_flagged_end_to_end(analyzer, base_config):
    cfg = {**base_config, "languages": ["de"]}
    text = "Der CAPEX und OPEX Aufwand wurde im Feld_Name erfasst; siehe RAG-Status."
    findings = detect_unit(analyzer, TextUnit("u1", text), cfg)
    flagged = {f.value for f in findings}
    for noise in ("CAPEX", "OPEX", "Feld_Name", "RAG"):
        assert noise not in flagged, f"{noise} should be filtered: {flagged}"


def test_number_plus_unit_not_an_address(analyzer, base_config):
    cfg = {**base_config, "languages": ["de"]}
    for text in ("Die Kosten betragen 66450 Euro.", "Aufwand 39870 Minuten insgesamt."):
        findings = detect_unit(analyzer, TextUnit("u1", text), cfg)
        assert not any(f.entity_type == "DE_ADDRESS" for f in findings), (text, [f.value for f in findings])
    # A real address still matches.
    findings = detect_unit(analyzer, TextUnit("u1", "Anschrift 80331 München im System."), cfg)
    assert any(f.entity_type == "DE_ADDRESS" for f in findings)


def test_english_name_anchors_catch_english_pii_in_german_scan(analyzer, base_config):
    """Layered mixed-language: an English name in a German-scanned document is
    still caught via the (language-independent) honorific/label anchors."""
    cfg = {**base_config, "languages": ["de"]}
    for text, name in [("Bitte an Mr Smith weiterleiten.", "Smith"),
                       ("Client: John Baker hat zugestimmt.", "John Baker")]:
        findings = detect_unit(analyzer, TextUnit("u1", text), cfg)
        assert any(f.entity_type == "PERSON" and name in f.value for f in findings), (text, [f.value for f in findings])


# --- Fifth wave: propagation bypassed the precision filters (2026-07-23),
# amplifying a bad seed (a snake_case field id, a common-word surname used as
# an ordinary word) across the whole document. ---


def test_propagation_findings_pass_the_same_precision_filters(analyzer, base_config):
    cfg = {**base_config, "languages": ["de"],
           "propagate": [("PERSON", "Aktueller_Status"), ("PERSON", "Gering"), ("PERSON", "Müller")]}
    # snake_case field id and a lowercase common-word occurrence must NOT propagate;
    # a real surname still does.
    text = "Feld Aktueller_Status geaendert, Beitrag ist gering; Herr Müller hat unterschrieben."
    findings = detect_unit(analyzer, TextUnit("u1", text), cfg)
    vals = {f.value for f in findings}
    assert "Aktueller_Status" not in vals, f"snake_case id propagated: {vals}"
    assert "gering" not in vals and "Gering" not in vals, f"lowercase common word propagated: {vals}"
    assert "Müller" in vals, f"real surname must still propagate: {vals}"


def test_looks_like_name_rejects_snake_case():
    from anonymizer.formats.xlsx_handler import _looks_like_name

    assert not _looks_like_name("Aktueller_Status")
    assert not _looks_like_name("Konzept_Gate_1_Status")
    assert _looks_like_name("Klaus Müller")


# --- Sixth wave: corroboration-only for ORG/LOCATION/MISC + jargon allow-list ---


def test_corroboration_only_drops_bare_org_loc_misc_guesses():
    from anonymizer import core
    from anonymizer.models import Finding as F

    cfg = {"entities": {}, "tiers": {"high": 0.9, "medium": 0.5}, "corroboration_only": True}
    findings = [
        F("ORGANIZATION", "OpenClaw", 0.85, "c", "u1", 0, 8, source="SpacyRecognizer"),   # bare guess -> drop
        F("LOCATION", "Bearbeitung", 0.85, "c", "u2", 0, 11, source="SpacyRecognizer"),    # bare guess -> drop
        F("PERSON", "Müller", 0.85, "c", "u3", 0, 6, source="SpacyRecognizer"),            # PERSON: kept
        # Authoritatively corroborated (a pattern/anchor recognizer, not a bare
        # NER guess and NOT mere propagation) -> kept.
        F("ORGANIZATION", "Signavio", 0.85, "c", "u4", 0, 8, source="PatternRecognizer"),
        # Propagation is DERIVED from NER, so a propagation-only ORG is still a
        # guess and is dropped.
        F("ORGANIZATION", "OpenClaw", 0.85, "c", "u5", 0, 8, source="propagation"),
    ]
    result = core.build_scan_result(findings, [TextUnit("u", "x")], cfg)
    vals = {g.value for g in result.all_actionable()}
    assert "OpenClaw" not in vals and "Bearbeitung" not in vals, f"bare/propagated ORG/LOC guesses must drop: {vals}"
    assert "Müller" in vals, "PERSON is never dropped by corroboration-only"
    assert "Signavio" in vals, "an authoritatively-corroborated ORG must survive"


def test_corroboration_only_off_keeps_everything():
    from anonymizer import core
    from anonymizer.models import Finding as F

    cfg = {"entities": {}, "tiers": {"high": 0.9, "medium": 0.5}, "corroboration_only": False}
    findings = [F("ORGANIZATION", "OpenClaw", 0.85, "c", "u1", 0, 8, source="SpacyRecognizer")]
    result = core.build_scan_result(findings, [TextUnit("u", "x")], cfg)
    assert any(g.value == "OpenClaw" for g in result.all_actionable())


def test_jargon_terms_in_shipped_allow_list():
    """Common business/tech jargon ships in the allow-list so NER never flags it."""
    import yaml

    from anonymizer.config import DEFAULT_CONFIG_PATH

    shipped = yaml.safe_load(DEFAULT_CONFIG_PATH.read_text(encoding="utf-8"))
    allow = set(shipped.get("allow_list", []))
    for term in ("CAPEX", "OPEX", "RAG", "Dashboard", "Lessons Learned"):
        assert term in allow, f"{term} should be in the shipped allow-list"


# --- Seventh wave: German nominalization nouns (Effizienz/Derivatefreiheit) ---


def test_german_nominalization_filter(analyzer, base_config):
    cfg = {**base_config, "languages": ["de"]}
    # Nominalizer-suffix common nouns tagged as entities -> filtered.
    for text in ("Effizienz", "Derivatefreiheit", "Die Nutzung ist gering.", "Reaktionszeiten sind hoch."):
        findings = detect_unit(analyzer, TextUnit("u", text), cfg)
        vals = {f.value for f in findings}
        assert not (vals & {"Effizienz", "Derivatefreiheit", "Nutzung", "Reaktionszeiten"}), (text, vals)


def test_nominalization_filter_spares_real_names(analyzer, base_config):
    """Ordinary surnames have no nominalizer suffix; short -ung surnames are
    spared by the length floor / PROPN check."""
    cfg = {**base_config, "languages": ["de"]}
    # Herr <Name> anchors it as a person so the assertion is about the filter,
    # not NER recall.
    for name in ("Müller", "Weber", "Bauer", "Metzler", "Jung"):
        findings = detect_unit(analyzer, TextUnit("u", f"Sehr geehrter Herr {name},"), cfg)
        assert any(name in f.value for f in findings), f"{name} was wrongly filtered"
