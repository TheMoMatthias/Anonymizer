"""Topical (non-personal) sensitivity — Phase A: header→category detection,
auto-gazetteer, category propagation, word-boundary header matching, corroboration
bypass, scan/apply parity. See docs/run_topical-sensitivity_2026-07-23.md.
"""

import openpyxl

from anonymizer.formats import xlsx_handler
from anonymizer.pipeline import apply_document, scan_document


def _topical_sheet(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Projektname", "Eingesetztes Tool", "Abteilung", "Lizenzgeber", "Beschreibung", "Produktgruppe"])
    ws.append(["Delphin", "Signavio", "Zahlungsverkehr", "OpenAI",
               "Wir nutzen Signavio zur Prozessmodellierung.", "Vorsorge"])
    ws.append(["Adler", "OpenClaw", "Kreditrisiko", "FactSet",
               "Das Tool OpenClaw ersetzt die Erfassung.", "Sparen"])
    p = tmp_path / "topical.xlsx"
    wb.save(p)
    return p


def test_header_to_category_detection(tmp_path, analyzer, base_config):
    result = scan_document(_topical_sheet(tmp_path), analyzer, base_config)
    by_type = {}
    for g in result.all_actionable():
        by_type.setdefault(g.entity_type, set()).add(g.value)
    assert "Signavio" in by_type.get("TOOL", set()) and "OpenClaw" in by_type.get("TOOL", set())
    assert by_type.get("DEPARTMENT", set()) >= {"Zahlungsverkehr", "Kreditrisiko"}
    assert by_type.get("LICENSEE", set()) >= {"OpenAI", "FactSet"}
    assert by_type.get("PROJECT", set()) >= {"Delphin", "Adler"}


def test_category_propagation_into_free_text(tmp_path, analyzer, base_config):
    """A tool named in a Tool column is also redacted where it recurs in a
    generic free-text cell (auto-gazetteer + propagation). Uses a non-category
    column ('Sonstiges') so the mention isn't already covered by a DESCRIPTION
    whole-cell summarize."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Eingesetztes Tool", "Sonstiges"])
    ws.append(["Signavio", "Migration erfolgt schrittweise."])
    ws.append(["OpenClaw", "Signavio wird auch hier eingesetzt."])
    p = tmp_path / "prop.xlsx"
    wb.save(p)
    result = scan_document(p, analyzer, base_config)
    tool = next(g for g in result.all_actionable() if g.entity_type == "TOOL" and g.value == "Signavio")
    assert tool.count >= 2, "Signavio should propagate from the Tool column into the free-text cell"


def test_word_boundary_header_not_oversensitive(tmp_path, analyzer, base_config):
    """'Produktgruppe' must NOT become a DEPARTMENT column via the term 'gruppe'
    (word-boundary matching), so its ordinary values aren't redacted."""
    result = scan_document(_topical_sheet(tmp_path), analyzer, base_config)
    vals = {g.value for g in result.all_actionable()}
    assert not (vals & {"Vorsorge", "Sparen"}), f"Produktgruppe values wrongly flagged: {vals}"


def test_topical_findings_survive_corroboration_only(tmp_path, analyzer, base_config):
    """Header/gazetteer matches are authoritative (guess=False) so corroboration-
    only never drops them."""
    result = scan_document(_topical_sheet(tmp_path), analyzer, base_config)
    topical = [g for g in result.all_actionable()
               if g.entity_type in ("TOOL", "DIVISION", "DEPARTMENT", "LICENSEE", "PROJECT")]
    assert topical and all(not g.is_ner_guess for g in topical)


def test_category_for_header_word_boundary():
    cfg = {"topical": {"enabled": True, "categories": {
        "TOOL": {"header_terms": ["tool", "system"]},
        "DEPARTMENT": {"header_terms": ["abteilung", "team"]},
    }}}
    assert xlsx_handler._category_for_header("Eingesetztes Tool", cfg) == "TOOL"
    assert xlsx_handler._category_for_header("Verantwortliche Abteilung", cfg) == "DEPARTMENT"
    assert xlsx_handler._category_for_header("Produktgruppe", cfg) is None
    assert xlsx_handler._category_for_header("Systematik", cfg) is None  # not a whole word
    assert xlsx_handler._category_for_header("", cfg) is None


def test_topical_disabled_detects_nothing(tmp_path, analyzer, base_config):
    cfg = {**base_config, "topical": {"enabled": False, "categories": base_config.get("topical", {}).get("categories", {})}}
    result = scan_document(_topical_sheet(tmp_path), analyzer, cfg)
    assert not any(g.entity_type in ("TOOL", "DIVISION", "DEPARTMENT", "LICENSEE", "PROJECT")
                   for g in result.all_actionable())


def test_topical_scan_apply_parity(tmp_path, analyzer, base_config, mapping_db_path):
    """apply must redact exactly what scan surfaced -- incl. the propagated tool
    in the description cell -- and the fail-loud verify must pass."""
    path = _topical_sheet(tmp_path)
    grouped = scan_document(path, analyzer, base_config).all_actionable()
    out_path, _report = apply_document(path, grouped, analyzer, base_config, mapping_db_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    # The description cell that named the tool must no longer contain it verbatim.
    body = "\n".join(str(c.value) for row in ws.iter_rows(min_row=2) for c in row if c.value)
    assert "Signavio" not in body and "OpenClaw" not in body, "propagated tool leaked into output"


# --- Phase B: summarize mode + DESCRIPTION category ---


def test_description_column_is_summarized(tmp_path, analyzer, base_config):
    """A free-text description column is summarized to a zero-content structural
    placeholder; the original prose must NOT survive into the output."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Projektname", "Projektbeschreibung"])
    ws.append(["Delphin", "Automatisierung der Kreditvergabe. Ziel ist schnellere Bearbeitung. Risiken bei Datenqualität."])
    p = tmp_path / "desc.xlsx"
    wb.save(p)

    result = scan_document(p, analyzer, base_config)
    desc = [g for g in result.all_actionable() if g.entity_type == "DESCRIPTION"]
    assert desc and desc[0].action == "summarize"
    # PROJECT name pseudonymizes (not summarized -- it's a name, not free text).
    proj = [g for g in result.all_actionable() if g.entity_type == "PROJECT"]
    assert proj and proj[0].action == "pseudonymize"


def test_summarize_placeholder_has_zero_original_content():
    from anonymizer.actions import resolve_replacement, _structural_summary
    from anonymizer.mapping import MappingStore
    import tempfile
    from pathlib import Path

    original = "Automatisierung der Kreditvergabe. Ziel ist schnellere Bearbeitung. Risiken bei Datenqualität."
    with tempfile.TemporaryDirectory() as tmp:
        out = resolve_replacement("DESCRIPTION", original, "summarize", MappingStore(Path(tmp) / "m.db"))
    assert out.startswith("[TEXT:") and out.endswith("]")
    # No run of original words survives (zero-content guarantee).
    for word in ("Automatisierung", "Kreditvergabe", "Datenqualität", "Bearbeitung"):
        assert word not in out, f"summary leaked original content: {out}"
    # Multi-sentence -> counted.
    assert "Sätze" in out


def test_summarize_survives_fail_loud_verify(tmp_path, analyzer, base_config, mapping_db_path):
    """A summarized description column must pass verify_output/_literal_residual
    (the placeholder contains no original text)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Beschreibung"])
    ws.append(["Vertrauliche Projektdetails zur neuen Plattform. Nicht teilen."])
    p = tmp_path / "verify.xlsx"
    wb.save(p)
    grouped = scan_document(p, analyzer, base_config).all_actionable()
    out_path, _ = apply_document(p, grouped, analyzer, base_config, mapping_db_path)  # raises if verify fails
    body = "\n".join(str(c.value) for row in openpyxl.load_workbook(out_path).active.iter_rows(min_row=2)
                     for c in row if c.value)
    assert "Vertrauliche" not in body and "Plattform" not in body
    assert "[TEXT:" in body


# --- Phase C: cell-level policy (Sheet!Coord) exception layer ---


def test_cell_policy_redacts_a_single_cell(tmp_path, analyzer, base_config, mapping_db_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tab"
    ws.append(["Wert"])  # non-category header, so only the cell policy acts
    ws.append(["Interne Details A"])
    ws.append(["Interne Details B"])
    p = tmp_path / "cells.xlsx"
    wb.save(p)

    grouped = scan_document(p, analyzer, base_config).all_actionable()
    cfg = {**base_config, "cell_policies": {"Tab!A2": "anonymize"}}
    out_path, _ = apply_document(p, grouped, analyzer, cfg, mapping_db_path)
    ws2 = openpyxl.load_workbook(out_path)["Tab"]
    assert ws2["A2"].value != "Interne Details A", "cell A2 should be redacted by its cell policy"
    assert ws2["A3"].value == "Interne Details B", "cell A3 (no policy) must be untouched"


def test_cell_policy_wins_over_column_policy(tmp_path, analyzer, base_config, mapping_db_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tab"
    ws.append(["Notiz"])
    ws.append(["Zelle zwei mit etwas Inhalt."])
    ws.append(["Zelle drei."])
    p = tmp_path / "cells2.xlsx"
    wb.save(p)

    grouped = scan_document(p, analyzer, base_config).all_actionable()
    # Whole column pseudonymize, but A2 overridden to summarize.
    cfg = {**base_config, "column_policies": {"Tab!A": "pseudonymize"},
           "cell_policies": {"Tab!A2": "summarize"}}
    out_path, _ = apply_document(p, grouped, analyzer, cfg, mapping_db_path)
    ws2 = openpyxl.load_workbook(out_path)["Tab"]
    assert str(ws2["A2"].value).startswith("["), "A2 replaced"
    assert "Zeichen" in str(ws2["A2"].value) or "Sätze" in str(ws2["A2"].value), "A2 should be SUMMARIZED (cell wins)"
    assert str(ws2["A3"].value).startswith("[") and "_" in str(ws2["A3"].value), "A3 should be pseudonymized (column policy)"


def test_cell_summary_lists_flagged_cells(tmp_path, analyzer, base_config):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tab"
    ws.append(["Name"])
    ws.append(["Klaus Müller"])
    p = tmp_path / "cs.xlsx"
    wb.save(p)
    result = scan_document(p, analyzer, base_config)
    assert result.cells, "flagged cells should be summarized for the Cells panel"
    keys = {c.key for c in result.cells}
    assert any(k.startswith("Tab!A") for k in keys)


# --- Manual gazetteer terms (Settings) + settings section render ---


def test_manual_topical_term_redacted_in_prose(tmp_path, analyzer, base_config):
    """A term the user adds to a category (a tool the model can't detect in
    prose, e.g. 'DeepL Pro') is redacted document-wide via propagation."""
    import copy
    cfg = copy.deepcopy(base_config)
    cfg.setdefault("topical", {}).setdefault("categories", {}).setdefault("TOOL", {"header_terms": [], "terms": []})
    cfg["topical"]["categories"]["TOOL"]["terms"] = ["DeepL Pro"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Sonstiges"])
    ws.append(["Wir prüfen die Möglichkeiten von DeepL Pro für Übersetzungen."])
    p = tmp_path / "manual.xlsx"
    wb.save(p)
    result = scan_document(p, analyzer, cfg)
    assert any(g.entity_type == "TOOL" and "DeepL Pro" in g.value for g in result.all_actionable())


def test_settings_topical_section_renders(base_config):
    from nicegui import Client, ui
    from nicegui.testing.general import nicegui_reset_globals, prepare_simulation

    from anonymizer.gui import settings_page

    prepare_simulation()
    with nicegui_reset_globals():
        @ui.page("/p")
        def _p():
            pass

        with Client(_p):
            settings_page._topical_section(base_config, [])
