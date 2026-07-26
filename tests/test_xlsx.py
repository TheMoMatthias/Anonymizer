import zipfile

import openpyxl

from anonymizer.formats import xlsx_handler
from anonymizer.pipeline import apply_document, scan_document

IBAN = "DE89370400440532013000"


def _raw_package_text(path) -> str:
    with zipfile.ZipFile(path) as zf:
        return "\n".join(zf.read(n).decode("utf-8", "ignore") for n in zf.namelist())


def test_detects_person_including_hidden_sheet(sample_xlsx, analyzer, base_config):
    grouped = scan_document(sample_xlsx, analyzer, base_config).all_actionable()
    assert any(g.entity_type == "PERSON" for g in grouped)
    units = xlsx_handler.extract_text_units(sample_xlsx)
    assert any("Hidden" in u.id for u in units)


def test_apply_replaces_cells(sample_xlsx, analyzer, base_config, mapping_db_path):
    grouped = scan_document(sample_xlsx, analyzer, base_config).all_actionable()
    for g in grouped:
        # Sheet NAMES are scanned now (they are a first-class leak -- see
        # test_xlsx_sheet_name_is_surfaced_and_redacted), and NER duly claims the
        # structural fixture names "Main"/"Hidden" as LOCATION / NER_MISC. A
        # reviewer skips those; removing them would rename the sheets, which is a
        # different test's subject.
        g.action = "skip" if g.value in ("Main", "Hidden") else "pseudonymize"
    out_path, report_path = apply_document(sample_xlsx, grouped, analyzer, base_config, mapping_db_path)

    assert out_path.suffix == ".xlsx"
    wb = openpyxl.load_workbook(out_path)
    assert wb["Main"]["A2"].value != "Hans Mueller"
    assert wb["Hidden"]["A2"].value != "Hans Mueller"


def test_xlsm_output_has_macros_stripped(tmp_path, analyzer, base_config, mapping_db_path):
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Name"
    wb.active["A2"] = "Hans Mueller"
    path = tmp_path / "sample.xlsm"
    wb.save(path)

    grouped = scan_document(path, analyzer, base_config).all_actionable()
    for g in grouped:
        g.action = "pseudonymize"
    out_path, _ = apply_document(path, grouped, analyzer, base_config, mapping_db_path)

    assert out_path.suffix == ".xlsx"


def test_xlsx_name_column_override_never_overlaps(analyzer, base_config):
    """Regression (CORRUPTION): the whole-cell name-column override was appended
    after overlap resolution, so it could partially overlap a finding NER did make
    (a KONTO number in the same cell) -> the cell splicer produced garbled tokens.
    The returned findings must be non-overlapping."""
    findings = xlsx_handler._analyze_cell_text(
        "Mueller, Konto 12345678", "Name", analyzer, {**base_config, "languages": ["de"]}
    )
    spans = sorted((f.start, f.end) for f in findings)
    for (s1, e1), (s2, e2) in zip(spans, spans[1:]):
        assert e1 <= s2, f"overlapping findings would corrupt the cell: {spans}"
    assert findings, "name-column cell yielded no findings at all"


def test_xlsx_header_straddling_span_is_clipped_not_dropped(analyzer, base_config):
    """Regression (LEAK): a finding whose span starts inside the injected 'header: '
    prefix but extends into the cell value was dropped wholesale, leaking the value.
    It must be clipped to the value side instead."""
    cfg = {**base_config, "languages": ["de"], "deny_list": ["Bemerkung: Geheimprojekt"]}
    findings = xlsx_handler._analyze_cell_text("Geheimprojekt", "Bemerkung", analyzer, cfg)
    assert any("Geheimprojekt" in f.value for f in findings), f"straddling value dropped: {findings}"


def test_xlsx_repeated_values_memoized_consistently(tmp_path, analyzer, base_config, mapping_db_path):
    """Detection/redaction is memoized per (header, cell-text) for speed on large
    'database' sheets. A value repeated across many cells and sheets must still be
    caught at EVERY occurrence, pseudonymize to the SAME token, and pass the
    fail-loud verify -- memoization must not drop or diverge any occurrence."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "A"
    ws["A1"] = "Name"
    ws2 = wb.create_sheet("B")
    ws2["A1"] = "Kunde"
    for r in range(2, 12):
        ws[f"A{r}"] = "Hans Mueller"   # 10 rows
        ws2[f"A{r}"] = "Hans Mueller"  # + 10 rows on another sheet = 20 occurrences
    path = tmp_path / "repeat.xlsx"
    wb.save(path)

    grouped = scan_document(path, analyzer, base_config).all_actionable()
    person_occurrences = sum(g.count for g in grouped if g.entity_type == "PERSON")
    assert person_occurrences >= 20, f"memoization dropped occurrences: {person_occurrences}"

    for g in grouped:
        # Only the people: NER also claims the one-letter sheet names, and removing
        # those would rename the sheets this test looks up by name.
        g.action = "pseudonymize" if g.entity_type == "PERSON" else "skip"
    out_path, _ = apply_document(path, grouped, analyzer, base_config, mapping_db_path)  # raises if verify fails
    out = openpyxl.load_workbook(out_path)
    tokens = {out["A"][f"A{r}"].value for r in range(2, 12)} | {out["B"][f"A{r}"].value for r in range(2, 12)}
    assert len(tokens) == 1, f"repeated value not consistently tokenized: {tokens}"
    assert next(iter(tokens)).startswith("[PERSON_"), f"unexpected token: {tokens}"


def test_xlsx_formula_string_literal_does_not_survive(tmp_path, analyzer, base_config, mapping_db_path):
    """FALSE-CLEAN (B1d): _cell_scan_text skips formula cells entirely, so PII
    written as a formula STRING LITERAL (="Kunde <IBAN>") was never scanned, never
    removed and never re-checked -- apply reported the file verified while the
    IBAN sat in xl/worksheets/sheet1.xml."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Hinweis"
    ws["A2"] = f'="Kunde {IBAN} bitte pruefen"'  # stored as a formula, data_type "f"
    path = tmp_path / "formula.xlsx"
    wb.save(path)

    grouped = scan_document(path, analyzer, base_config).all_actionable()
    assert any(IBAN in g.value.replace(" ", "") for g in grouped), "formula literal must be surfaced for review"
    for g in grouped:
        g.action = "anonymize"
    out_path, _ = apply_document(path, grouped, analyzer, base_config, mapping_db_path)

    assert IBAN not in _raw_package_text(out_path), "formula string literal leaked the IBAN"
    # The formula itself must survive as a formula -- redaction replaces the PII
    # inside the quoted literal, it does not corrupt the expression.
    # Looked up by POSITION, not by name: the default title "Sheet" is itself a
    # decided value here, so the sheet is legitimately renamed (see B3).
    out_cell = openpyxl.load_workbook(out_path).worksheets[0]["A2"]
    assert isinstance(out_cell.value, str) and out_cell.value.startswith("=")


def test_xlsx_hyperlink_target_does_not_survive(tmp_path, analyzer, base_config, mapping_db_path):
    """FALSE-CLEAN (B1b): a cell hyperlink's TARGET is a relationship attribute,
    not cell text, so openpyxl's cell walk never scanned it and the handler-based
    re-scan never saw it. xlsx_handler.scan builds its own unit stream, so the
    auxiliary surface has to be wired into that path separately from
    extract_text_units -- this test covers exactly that."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Kontakt"
    ws["A2"] = "Ansprechpartner"
    ws["A2"].hyperlink = "mailto:hans.mueller@bank.de"
    path = tmp_path / "link.xlsx"
    wb.save(path)

    grouped = scan_document(path, analyzer, base_config).all_actionable()
    assert any("hans.mueller@bank.de" in g.value for g in grouped), "hyperlink target must be surfaced for review"
    for g in grouped:
        g.action = "anonymize"
    out_path, _ = apply_document(path, grouped, analyzer, base_config, mapping_db_path)

    assert "hans.mueller@bank.de" not in _raw_package_text(out_path), "cell hyperlink target leaked"


def test_xlsx_document_title_is_scrubbed(tmp_path, analyzer, base_config, mapping_db_path):
    """FALSE-CLEAN (B3): the workbook's docProps/core.xml title is not body text,
    so a customer name there was never scanned nor re-checked."""
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Umsatzliste"
    wb.properties.title = "Kreditakte Hans Mueller"
    wb.properties.description = "Erstellt fuer Hans Mueller"
    path = tmp_path / "titled.xlsx"
    wb.save(path)

    grouped = scan_document(path, analyzer, base_config).all_actionable()
    for g in grouped:
        g.action = "anonymize"
    out_path, _ = apply_document(path, grouped, analyzer, base_config, mapping_db_path)

    assert "Hans Mueller" not in _raw_package_text(out_path), "workbook core.xml title leaked the name"


def test_column_summary_lists_headers_and_counts(analyzer, base_config, tmp_path):
    """Column summary reports each column's header + how many findings landed in it,
    so the reviewer can set a whole-column policy without opening the file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Kunde", "Projekt", "Notiz"])
    ws.append(["Hans Mueller", "Geheimprojekt Nordwind", "offen"])
    ws.append(["Petra Weber", "Fusion", "zu"])
    path = tmp_path / "cols.xlsx"
    wb.save(path)

    result = scan_document(path, analyzer, {**base_config, "languages": ["de"]})
    cols = {c.key: c for c in result.columns}
    assert cols["Data!A"].header == "Kunde" and cols["Data!A"].pii_count >= 2  # two names
    assert cols["Data!B"].header == "Projekt"
    assert cols["Data!C"].header == "Notiz"


def test_column_blackout_redacts_undetected_cells_and_verifies(analyzer, base_config, tmp_path, mapping_db_path):
    """A whole-column blackout redacts EVERY non-empty cell -- including a topic-
    sensitive cell entity detection can't judge -- tokenizes repeats consistently,
    leaves empty cells alone, and still passes the fail-loud verify."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Kunde", "Projekt"])
    ws.append(["Hans Mueller", "Streng geheimes Vorhaben Nordwind"])
    ws.append(["Petra Weber", "Marktstrategie 2027"])
    ws.append(["Klaus Bauer", "Streng geheimes Vorhaben Nordwind"])  # repeat -> same token
    ws.append(["Anna Klein", ""])  # empty -> stays empty
    path = tmp_path / "blackout.xlsx"
    wb.save(path)

    cfg = {**base_config, "languages": ["de"]}
    grouped = scan_document(path, analyzer, cfg).all_actionable()
    for g in grouped:
        g.action = "pseudonymize"
    apply_cfg = {**cfg, "column_policies": {"Data!B": "pseudonymize"}}
    out_path, _ = apply_document(path, grouped, analyzer, apply_cfg, mapping_db_path)  # raises if verify fails

    out = openpyxl.load_workbook(out_path)["Data"]
    assert out["B2"].value.startswith("[PROJEKT_"), f"undetected cell not blacked out: {out['B2'].value!r}"
    assert out["B2"].value == out["B4"].value, "repeated column value not consistently tokenized"
    assert out["B5"].value in (None, ""), "empty cell must stay empty"
    assert out["A2"].value != "Hans Mueller", "name column must still be redacted via the value path"


def test_bare_ner_guess_does_not_rename_a_sheet(tmp_path, analyzer, base_config, mapping_db_path):
    """A sheet TITLE is structural: renaming it rewrites xl/workbook.xml and every
    formula, defined name and chart reference that points at the sheet.

    Measured: spaCy claims a bare "Tab" as PERSON at its flat 0.85. PERSON is
    deliberately exempt from the corroboration-only rule AND from the
    lowercase/stopword precision filters -- both correct, so a real lowercase
    surname stays reachable in a CELL -- which left nothing at all between one weak
    three-character guess and a rewritten workbook structure."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tab"
    ws.append(["Wert"])
    ws.append(["Interne Notiz"])
    path = tmp_path / "tab.xlsx"
    wb.save(path)

    grouped = scan_document(path, analyzer, base_config).all_actionable()
    assert not any(g.value == "Tab" for g in grouped), (
        f"a bare NER guess on a sheet TITLE must not become a finding: "
        f"{[(g.entity_type, g.value) for g in grouped]}"
    )
    for g in grouped:
        g.action = "anonymize"
    out_path, _ = apply_document(path, grouped, analyzer, base_config, mapping_db_path)
    assert "Tab" in openpyxl.load_workbook(out_path).sheetnames, (
        "the sheet was renamed on the strength of a bare NER guess"
    )


def test_pattern_backed_sheet_title_is_still_renamed(tmp_path, analyzer, base_config, mapping_db_path):
    """The corroboration gate must not weaken the leak it guards. An IBAN in a
    sheet title is pattern-anchored and checksum-tested -- not an NER guess at all
    -- so it still renames the sheet and must not survive in xl/workbook.xml."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Konto {IBAN}"[:31]
    ws.append(["Wert"])
    ws.append(["Zahlungseingang"])
    path = tmp_path / "kontosheet.xlsx"
    wb.save(path)

    grouped = scan_document(path, analyzer, base_config).all_actionable()
    assert any(IBAN in g.value for g in grouped), "a pattern-backed sheet title must still be surfaced"
    for g in grouped:
        g.action = "anonymize"
    out_path, _ = apply_document(path, grouped, analyzer, base_config, mapping_db_path)
    assert IBAN not in _raw_package_text(out_path), "the IBAN leaked in the sheet title"


def test_xlsx_sheet_name_is_surfaced_and_redacted(tmp_path, analyzer, base_config, mapping_db_path):
    """FALSE-CLEAN (B3): Excel SHEET NAMES were never scanned and never redacted.
    The authoritative copy is xl/workbook.xml <sheet name="...">; only the app.xml
    TitlesOfParts CACHE was blanked, which removed the evidence and left the leak.
    For this deployment a workbook with ONE SHEET PER CLIENT is a normal shape, so
    the customer's name sat in the tab of a file marked verified. Renaming must not
    break the formulas and defined names that reference the sheet."""
    from openpyxl.workbook.defined_name import DefinedName

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kunde Hans Mueller"
    ws["A1"] = "Umsatz"
    ws["A2"] = 4200
    other = wb.create_sheet("Uebersicht")
    other["A1"] = "Summe"
    other["A2"] = "='Kunde Hans Mueller'!A2"
    wb.defined_names.add(DefinedName("Kundensumme", attr_text="'Kunde Hans Mueller'!$A$2"))
    path = tmp_path / "persheet.xlsx"
    wb.save(path)

    units = xlsx_handler.extract_text_units(path)
    assert any(u.text == "Kunde Hans Mueller" for u in units), "sheet name must be a scannable text unit"

    grouped = scan_document(path, analyzer, base_config).all_actionable()
    assert any("Hans Mueller" in g.value for g in grouped), "sheet-name PII must be surfaced for review"
    for g in grouped:
        g.action = "pseudonymize"
    out_path, _ = apply_document(path, grouped, analyzer, base_config, mapping_db_path)  # raises if verify fails

    assert "Hans Mueller" not in _raw_package_text(out_path), "sheet name leaked into xl/workbook.xml"
    out = openpyxl.load_workbook(out_path)
    assert "Kunde Hans Mueller" not in out.sheetnames
    new_title = next(t for t in out.sheetnames if t != "Uebersicht")
    # The reference must have followed the rename -- a dangling '#REF' formula is a
    # corrupted deliverable, which is why the rename is verified rather than hoped.
    assert new_title in out["Uebersicht"]["A2"].value, out["Uebersicht"]["A2"].value
    assert new_title in out.defined_names["Kundensumme"].value


def test_column_entity_type_readable_and_safe():
    assert xlsx_handler._column_entity_type("Projekt", "B") == "PROJEKT"
    assert xlsx_handler._column_entity_type("Kunden-Nr.", "A") == "KUNDEN_NR"
    assert xlsx_handler._column_entity_type("", "D") == "COLUMN_D"  # no header -> column letter


def test_column_entity_type_transliterates_umlauts(mapping_db_path):
    """UNREVERSIBLE OUTPUT (P6): the label kept German umlauts, but actions.TOKEN_RE
    is [A-Z0-9_] and can never match one -- so a pseudonymized 'Pruefung' column
    written as [PRUEFUNG_1] with an umlaut could NEVER be re-identified, silently
    and permanently. Transliterate instead, so the label stays readable AND round-
    trips through reidentify_text."""
    from anonymizer.actions import reidentify_text
    from anonymizer.mapping import MappingStore

    assert xlsx_handler._column_entity_type("Prüfung", "B") == "PRUEFUNG"
    assert xlsx_handler._column_entity_type("Größe", "C") == "GROESSE"
    assert xlsx_handler._column_entity_type("Änderung Öl Übertrag", "D") == "AENDERUNG_OEL_UEBERTRAG"

    with MappingStore(mapping_db_path) as store:
        entity = xlsx_handler._column_entity_type("Prüfung", "B")
        token = f"[{store.get_or_create(entity, 'geheim', label=entity)}]"
        store.save()
        restored, n = reidentify_text(f"Zelle: {token}", store)
    assert n == 1 and "geheim" in restored, f"column token is not reversible: {token}"


def test_name_header_re_widened_and_configurable():
    """The built-in people-header set now covers common German business headers,
    is extendable via config, and does not match non-people headers."""
    assert xlsx_handler._name_header_re().search("Projektleiter")  # widened built-in
    assert xlsx_handler._name_header_re().search("Betreuer")
    assert xlsx_handler._name_header_re().search("Verantwortlich")
    assert not xlsx_handler._name_header_re().search("Betrag")  # not a people column
    assert not xlsx_handler._name_header_re().search("Sachwalter")  # only via config...
    assert xlsx_handler._name_header_re(("Sachwalter",)).search("Sachwalter")  # ...added here


def test_xlsx_configured_name_header_claims_bare_surname(analyzer, base_config):
    """A workbook-specific header added via config['name_column_headers'] makes the
    whole cell a person -- catching a bare common-noun surname NER misses in a cell."""
    cfg = {**base_config, "languages": ["de"], "name_column_headers": ["Sachwalter"]}
    findings = xlsx_handler._analyze_cell_text("Weber", "Sachwalter", analyzer, cfg)
    assert any("Weber" in f.value for f in findings), f"configured header did not claim the cell: {findings}"
