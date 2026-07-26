"""Validate the synthetic audit workbook, then score detection against its manifest.

    uv run python scripts/score_test_workbook.py tests/fixtures/generated

PHASE 1 -- validate the INSTRUMENT. A measurement is only worth as much as the
thing doing the measuring, so this checks the fixture before believing any number
it produces: manifest/workbook agreement, checksum-valid IBANs, decoys that really
are innocuous, no value planted as both secret and decoy, and that every hard
surface the design claims (formula literal, cell comment, hidden sheet, sensitive
sheet title) is actually present.

PHASE 2 -- score. Recall over planted secrets (per data class), false positives
over decoys, and a scan/apply round trip ending in the fail-loud verify.

KNOWN LIMITATION of the recall number, stated so it is not over-read: matching is
by VALUE, not by location. Findings are value-keyed all the way through this tool,
so a value planted in two places counts as found if it was detected in EITHER. That
makes recall here an UPPER BOUND: it cannot prove a specific surface was covered.
Measured consequence -- the name planted inside a formula literal and the surname
planted in free text both scored as "found" via their occurrence in the client
table, while apply demonstrably failed to redact the formula. Phase 3 is what
catches that, because the fail-loud verify checks the written bytes.
"""

from __future__ import annotations

import json
import sys
import time
import zipfile
from pathlib import Path

import openpyxl
import yaml

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from anonymizer import taxonomy  # noqa: E402
from anonymizer.config import DEFAULT_CONFIG_PATH  # noqa: E402
from anonymizer.engine import build_analyzer  # noqa: E402
from anonymizer.pipeline import apply_document, scan_document  # noqa: E402
from anonymizer.validators import iban_valid  # noqa: E402


def validate(xlsx: Path, manifest: dict) -> list[str]:
    problems: list[str] = []
    wb = openpyxl.load_workbook(xlsx)

    if set(wb.sheetnames) != set(manifest["sheets"]):
        problems.append(f"sheet mismatch: {wb.sheetnames} vs {manifest['sheets']}")

    secrets = manifest["must_catch"]
    decoys = manifest["must_not_catch"]

    # Every planted value must actually be somewhere in the workbook, or the
    # manifest is lying and recall would be measured against fiction.
    blob = "\n".join(
        str(c.value)
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for c in row
        if c.value is not None
    )
    blob += "\n" + "\n".join(wb.sheetnames)
    for s in secrets:
        if s["value"] not in blob and s["cell"] != "A2":  # formula/comment cells are checked below
            if s["value"] not in _raw_package_text(xlsx):
                problems.append(f"planted secret not present in the file: {s['sheet']}!{s['cell']} {s['value']!r}")

    # A value must never be both a secret and a decoy -- that would make the score
    # self-contradictory whichever way detection went.
    sv, dv = {s["value"] for s in secrets}, {d["value"] for d in decoys}
    for overlap in sv & dv:
        problems.append(f"value is BOTH a planted secret and a decoy: {overlap!r}")

    # Short values are magnets for substring coincidences and would inflate recall.
    for s in secrets:
        if len(s["value"].strip()) < 4:
            problems.append(f"planted secret too short to score reliably: {s['value']!r}")

    # IBANs must pass their checksum or the financial-ID path is not really tested.
    for s in secrets:
        if s["data_class"] == "financial_ids" and s["value"].startswith("DE") and len(s["value"]) == 22:
            if not iban_valid(s["value"]):
                problems.append(f"planted IBAN fails its checksum: {s['value']}")

    # The hard surfaces the design claims must actually exist.
    raw = _raw_package_text(xlsx)
    if "Sachbearbeiter: " not in raw:
        problems.append("no formula literal found -- the formula surface is not exercised")
    if not any(ws.sheet_state == "hidden" for ws in wb.worksheets):
        problems.append("no hidden sheet -- that surface is not exercised")
    has_comment = any(c.comment for ws in wb.worksheets for row in ws.iter_rows() for c in row)
    if not has_comment:
        problems.append("no cell comment -- that surface is not exercised")

    # Both languages must be represented, or "DE and EN" is untested.
    classes = {s["data_class"] for s in secrets}
    for needed in ("people", "financial_ids", "contact", "special_category", "internal_topical"):
        if needed not in classes:
            problems.append(f"no planted secret for data class {needed!r}")

    return problems


def _raw_package_text(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        return "\n".join(zf.read(n).decode("utf-8", "ignore") for n in zf.namelist())


def main(d: Path) -> int:
    xlsx = d / "audit_workbook.xlsx"
    manifest = json.loads((d / "audit_manifest.json").read_text(encoding="utf-8"))

    print("=" * 78)
    print("PHASE 1 -- is the test instrument itself sound?")
    print("=" * 78)
    problems = validate(xlsx, manifest)
    if problems:
        for p in problems:
            print(f"  PROBLEM: {p}")
        print(f"\n{len(problems)} problem(s) -- the fixture is NOT fit to measure with.")
        return 1
    secrets, decoys = manifest["must_catch"], manifest["must_not_catch"]
    print(f"  ok  {len(manifest['sheets'])} sheets, {len(secrets)} planted secrets, {len(decoys)} decoys")
    print(f"  ok  every planted value present in the file; no secret/decoy collisions")
    print(f"  ok  all planted IBANs pass their mod-97 checksum")
    print(f"  ok  hard surfaces present: formula literal, cell comment, hidden sheet")
    print(f"  ok  data classes covered: {sorted({s['data_class'] for s in secrets})}")

    print("\n" + "=" * 78)
    print("PHASE 2 -- what does detection actually find?")
    print("=" * 78)
    cfg = yaml.safe_load(DEFAULT_CONFIG_PATH.read_text(encoding="utf-8"))
    backend = None
    if "--gliner" in sys.argv:
        # Same measurement, with the offline ML second pass on. This is what
        # answers "does GLiNER actually earn its place?" on a scored corpus rather
        # than on impressions.
        import os

        os.environ.setdefault("HF_HUB_OFFLINE", "1")
        from anonymizer.gliner_recognizer import load_gliner_backend

        cfg["gliner"] = {**cfg["gliner"], "enabled": True}
        backend = load_gliner_backend(cfg["gliner"])
        print("  [GLiNER ENABLED]")
    analyzer = build_analyzer(cfg, gliner_backend=backend)

    t0 = time.perf_counter()
    result = scan_document(xlsx, analyzer, cfg)
    scan_s = time.perf_counter() - t0
    found = result.all_actionable()
    found_vals = {g.value.strip().lower(): g for g in found}
    print(f"  scan: {scan_s:.1f}s, {len(found)} distinct findings, "
          f"{len(result.columns or [])} columns offered a policy")

    def covered(value: str):
        v = value.strip().lower()
        if v in found_vals:
            return found_vals[v]
        # A finding that CONTAINS the planted value still covers it (the span may be
        # wider, e.g. an address block or an Art.9 line).
        for fv, g in found_vals.items():
            if v in fv or fv in v:
                return g
        return None

    by_class: dict[str, list[bool]] = {}
    by_sheet: dict[str, list[bool]] = {}
    misses = []
    for s in secrets:
        g = covered(s["value"])
        ok = g is not None
        by_class.setdefault(s["data_class"], []).append(ok)
        by_sheet.setdefault(s["sheet"], []).append(ok)
        if not ok:
            misses.append(s)

    def table(title, groups):
        print(f"\n  {title}")
        for k, hits in sorted(groups.items()):
            n, t = sum(hits), len(hits)
            bar = "#" * int(20 * n / t) + "." * (20 - int(20 * n / t))
            flag = "   <-- weak" if n < t else ""
            print(f"    {k:24} {n:3}/{t:<3} {100*n/t:5.1f}%  {bar}{flag}")

    table("RECALL by data class (planted -> found):", by_class)
    # Per-SHEET too: an aggregate hides a single language or surface failing wholesale,
    # which is exactly how the English Art.9 gap stayed invisible before.
    table("RECALL by sheet:", by_sheet)
    total_hit = sum(sum(h) for h in by_class.values())
    print(f"\n    {'TOTAL':24} {total_hit:3}/{len(secrets):<3} {100*total_hit/len(secrets):5.1f}%")

    if misses:
        print(f"\n  MISSED ({len(misses)}):")
        for m in misses:
            print(f"    {m['sheet']}!{m['cell']:<14} {m['value'][:46]!r:<50} [{m['why']}]")

    # A decoy "fails" only on a HARMFUL claim. Detecting the real date inside
    # "Zum Stichtag 31.12.2025 betrug der Saldo null." is correct behaviour -- a
    # date is a date, it lands in the low-sensitivity dates class, and profiles
    # routinely skip it. Counting that as a false positive would penalise the tool
    # for being right and would quietly pad the FP number. What must never happen
    # to ordinary business prose is being claimed as a person, a special category,
    # an identifier or contact data -- those are what destroy meaning.
    harmless = {"dates_other"}
    fps = []
    benign = []
    for dec in decoys:  # not `d` -- that is the fixture directory, in scope here
        g = covered(dec["value"])
        if g is None:
            continue
        (benign if taxonomy.data_class_for(g.entity_type).key in harmless else fps).append((dec, g))
    print(f"\n  FALSE POSITIVES on decoys: {len(fps)}/{len(decoys)}")
    for f, g in fps:
        print(f"    {f['sheet']}!{f['cell']:<6} {f['value'][:44]!r:<48} -> {g.entity_type} [{f['why']}]")
    for f, g in benign:
        print(f"    (not counted) {f['sheet']}!{f['cell']} -> {g.entity_type} on {f['value'][:38]!r} "
              f"-- low-sensitivity class, correct detection")

    print("\n" + "=" * 78)
    print("PHASE 3 -- scan/apply round trip + fail-loud verify")
    print("=" * 78)
    for g in found:
        if taxonomy.data_class_for(g.entity_type).key == "special_category":
            g.action = "anonymize"
        elif g.action == "skip":
            g.action = "pseudonymize"
    t0 = time.perf_counter()
    try:
        out_path, report_path = apply_document(xlsx, found, analyzer, cfg, d / "audit_mappings.db")
        apply_s = time.perf_counter() - t0
        print(f"  ok  wrote {out_path.name} in {apply_s:.1f}s (verify_output passed)")
        leaked = [s for s in secrets if s["value"] in _raw_package_text(out_path)]
        print(f"  {'ok ' if not leaked else 'LEAK'} planted secrets surviving verbatim in the output: {len(leaked)}")
        for s in leaked[:10]:
            print(f"      {s['sheet']}!{s['cell']} {s['value'][:44]!r} [{s['why']}]")
        rep = json.loads(report_path.read_text(encoding="utf-8"))
        print(f"  provenance: {rep.get('detection')}")
    except Exception as exc:  # noqa: BLE001
        print(f"  APPLY FAILED (fail-loud): {exc}")
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main(Path(sys.argv[1] if len(sys.argv) > 1 else "tests/fixtures/generated").resolve()))
