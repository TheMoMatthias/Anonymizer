"""Generate a synthetic bank workbook + a GROUND-TRUTH manifest for auditing detection.

    uv run python scripts/make_test_workbook.py tests/fixtures/generated

Writes `audit_workbook.xlsx` and `audit_manifest.json`. The manifest is what makes
this a measurement instrument rather than a demo file: it records every planted
secret (value, sheet, cell, expected data class) AND every deliberate DECOY that
must survive untouched. `scripts/score_test_workbook.py` reads it back and reports
recall and false-positive rates per data class and per sheet.

Everything here is FICTIONAL -- invented names, invented projects, invented
counterparties, invented figures, invented hostnames. Nothing derives from a real
customer and no real file is read.

That is a hard rule, not a nicety: this fixture is the measurement instrument, and
an instrument built out of real customer data cannot live in a git repo, be copied
to a build machine, or be attached to a bug report. Where a real file has been
studied (the innovation-pipeline workbook audited on 2026-07-27, see
docs/run_precision-rework_2026-07-27.md), only its STRUCTURE was reproduced --
sheet layout, column names, value style, language mix, length distributions. Every
value was written from scratch. Nothing was copied, and nothing was "anonymized by
shuffling", which would have left the real strings in the repo permanently.

DESIGN: realistic first, but deliberately containing the shapes that break
detection. Cell-level lists are the easy case; the hard case is PROSE, so most of
the volume here is sentences of the kind a bank actually writes:

  * several people in ONE sentence, some with titles (Herr/Frau/Dr./Ms/Mr)
  * bare surnames with no first name and no sentence to lean on
  * German genitives ("Muellers Team", "Webers Bereich")
  * surnames that are also ordinary German words (Bauer, Weber, Koch, Schneider)
  * a person whose surname is a city (Frankfurt), and two people sharing a surname
  * compound/hyphenated surnames (Meyer-Landrut, Schulz-Baumgarten)
  * cost projections in five different formats (EUR 4,2 Mio. / 1.250.000 EUR /
    TEUR 850 / EUR 2.4m / USD 3,100,000)
  * project descriptions, decisions and meeting minutes in German AND English
  * Art. 9 data in both languages, in prose and in labelled fields

Plus the structural surfaces a cell walk misses: a literal inside a FORMULA, a
cell COMMENT, a HIDDEN sheet, and a sensitive SHEET NAME.

The seed is fixed so the file is reproducible: an audit you cannot re-run is an
anecdote. Pass a different seed as argv[2] to generate an independent variant.
"""

from __future__ import annotations

import json
import random
import sys
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

SEED = 20260726

# --- fictional content pools -------------------------------------------------
FIRST_DE = ["Klaus", "Petra", "Anja", "Thomas", "Miriam", "Jonas", "Elif", "Lukas", "Sofia", "Bernd",
            "Katharina", "Matthias", "Yasemin", "Fabian"]
# Deliberately loaded with surnames that are also ordinary German nouns/adjectives
# (Bauer=farmer, Weber=weaver, Koch=cook, Schneider=tailor, Fischer=fisher,
# Richter=judge, Jung=young) -- the exact shape the precision filters must not eat.
LAST_DE = ["Mueller", "Weber", "Bauer", "Schneider", "Hoffmann", "Koch", "Fischer", "Richter",
           "Jung", "Yilmaz", "Kowalski", "Brandt", "Meyer-Landrut", "Schulz-Baumgarten", "Frankfurt"]
FIRST_EN = ["Sarah", "Liam", "Amina", "Oliver", "Priya", "Daniel", "Grace", "Noah"]
LAST_EN = ["Okonkwo", "Fitzgerald", "Haddad", "O'Brien", "Nakamura", "Whitfield", "Adeyemi"]

CITIES = [("50667", "Koeln"), ("60311", "Frankfurt am Main"), ("20095", "Hamburg"),
          ("80331", "Muenchen"), ("70173", "Stuttgart")]
STREETS = ["Hauptstrasse", "Kirchweg", "Rudolf-Breitscheid-Str.", "Am Wall", "Lindenallee",
           "Zum Alten Forsthaus", "Koenigsallee"]
PROJECTS = ["Delphin", "Nordstern", "Kolibri", "Adler", "Seidenpfad", "Marschall", "Habicht", "Zeitreise"]
TOOLS = ["Signavio", "DeepL Pro", "Camunda", "Alteryx", "Collibra", "OpenClaw", "Murex", "Kondor+"]
DEPTS = ["Zahlungsverkehr", "Kreditrisiko", "Treasury", "Compliance", "Marktfolge", "Meldewesen"]
DIVISIONS = ["Firmenkunden", "Privatkunden", "Kapitalmarkt", "Vermoegensverwaltung"]
VENDORS = ["Aurexa Systems GmbH", "Northgate Analytics Ltd", "Vertano AG", "Blauwerk Software",
           "Helios Data Partners", "Cormont Consulting"]
DIAGNOSES = ["Diabetes mellitus Typ 2, insulinpflichtig", "chronische Migraene",
             "Bandscheibenvorfall L4/L5", "Burnout-Syndrom, stationaer behandelt"]
CONFESSIONS = ["roemisch-katholisch", "evangelisch", "konfessionslos", "muslimisch"]
UNIONS = ["ver.di", "IG Metall", "Marburger Bund", "Vereinigung Cockpit"]
COSTS = ["EUR 4,2 Mio.", "1.250.000 EUR", "TEUR 850", "EUR 2.4m", "USD 3,100,000",
         "EUR 640.000", "TEUR 1.900"]

# --- the "database workbook" archetype (sheets 12-16) -------------------------
# A second, structurally different shape from the bank-letter sheets above: an
# internal innovation-pipeline tracker, one wide DB_* sheet per phase, with
# `_`-joined schema-style headers. Modelled on the STRUCTURE of a real internal
# workbook whose export was audited on 2026-07-27; every VALUE here is invented.
# Nothing is copied, shuffled or derived from that file -- see the module docstring.
#
# This archetype is what actually broke detection, and in ways the sheets above
# cannot reach:
#   * people in ENGLISH-headed columns (Owner / MDX_Lead / User), where the German
#     header stems never matched and bare spaCy misses a lone first name in a cell
#   * internal LINKS carrying a hostname plus a document id
#   * enumerated dropdown values repeated hundreds of times, which dominated the
#     false positives -- declared as real Excel data validations so the workbook's
#     own controlled vocabulary is readable
#   * German compound nouns in free text, the residual noise a POS gate cannot reject

# Bare first names, as an MDX_Lead / MDX_Proxy column actually holds them. A lone
# capitalized first name in a cell is the single hardest people shape: no
# honorific, no surname, no sentence.
LEADS = ["Siggi", "Cordula", "Mirijam", "Hendrik", "Marco", "Sergii", "Anneke", "Jorick"]
# Invented innovation-pipeline project titles (German/English mix, as such a
# workbook really is).
INNO_TITLES = [
    "Automatisierte Belegpruefung", "Agentischer RFP-Assistent",
    "Knowledge Graph fuer Fondsdaten", "Self-Service Reporting Portal",
    "Digitale Depoteroeffnung", "Sanctions Screening Copilot",
    "Stammdaten-Harmonisierung", "Meeting Minutes Generator",
]
# Fictional corporate hosts. `beispielbank.de` is invented ("example bank"); the
# link SHAPE is what matters -- a hostname that names an internal system plus a
# path that is a direct document pointer.
INNO_HOSTS = ["confluence.beispielbank.de", "dms.beispielbank.de", "prozesse.beispielbank.de"]
# Controlled vocabulary: the values a dropdown column repeats on every row. These
# are DECOYS -- a scale label is never personal data, however often a name model
# mistakes one for a surname.
SCALE_3PKT = ["0 - Kein Beitrag", "1 - Gering / Kaum", "2 - Mittel / Bedingt", "3 - Hoch / Signifikant"]
STATUS_LIST = ["Ungeprueft", "In Pruefung", "Abgeschlossen", "Pausiert", "Ausstehend", "Geklaert"]
RELEVANZ_LIST = ["Idee", "Validierung", "Konzeption", "Rollout"]
PERSONA_ROLES = ["Hauptzielgruppe", "Nebenzielgruppe", "Nur indirekt betroffen"]
# German compound nouns and business jargon that a German NER model routinely tags
# as a PERSON at its flat score. Capitalized (German capitalizes every noun) and
# noun-class, so neither the case filter nor the POS filter can reject them --
# which is precisely why they are planted as decoys rather than assumed harmless.
JARGON_DE = [
    "Datenfeeds", "Kernworkflow", "Portfoliobeitrag", "Abrechnungsprozess",
    "Zielnutzer", "Marktdatengrundlage", "Content-Framework", "Datenlayer",
    "Standardwechsel", "Anbindung", "Bearbeitungszeit", "Folgeaufwand",
]


def main(out_dir: Path, seed: int = SEED) -> int:
    rng = random.Random(seed)
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    must_catch: list[dict] = []
    must_not_catch: list[dict] = []

    def secret(sheet, cell, value, cls, why):
        must_catch.append({"sheet": sheet, "cell": cell, "value": value, "data_class": cls, "why": why})
        return value

    def decoy(sheet, cell, value, why):
        must_not_catch.append({"sheet": sheet, "cell": cell, "value": value, "why": why})
        return value

    def iban(n: int) -> str:
        bban = f"{37040044:08d}{n:010d}"
        check = 98 - (int(bban + "131400") % 97)
        return f"DE{check:02d}{bban}"

    de_people = [(rng.choice(FIRST_DE), rng.choice(LAST_DE)) for _ in range(12)]
    # Two people deliberately share a surname -- a bare "Hoffmann" then belongs to
    # neither unambiguously, which is exactly the propagation trade-off we accepted.
    de_people[3] = (de_people[3][0], de_people[2][1])
    en_people = [(rng.choice(FIRST_EN), rng.choice(LAST_EN)) for _ in range(6)]

    # ---------------------------------------------------------------- 1. Kundenstamm (DE)
    ws = wb.active
    ws.title = "Kundenstamm"
    ws.append(["KundenNr", "Nachname", "Vorname", "IBAN", "Telefon", "Anschrift", "Geburtsdatum"])
    for i, (fn, ln) in enumerate(de_people, start=2):
        plz, city = rng.choice(CITIES)
        addr = f"{rng.choice(STREETS)} {rng.randint(1, 99)}, {plz} {city}"
        ib = iban(rng.randint(10**8, 10**9))
        phone = f"0{rng.choice([170, 171, 151, 160])} {rng.randint(1000000, 9999999)}"
        dob = f"{rng.randint(1,28):02d}.{rng.randint(1,12):02d}.{rng.randint(1955,1999)}"
        ws.append([f"K-{rng.randint(10000,99999)}", ln, fn, ib, phone, addr, dob])
        secret("Kundenstamm", f"B{i}", ln, "people", "bare surname in a name column")
        secret("Kundenstamm", f"D{i}", ib, "financial_ids", "IBAN, checksum-valid")
        secret("Kundenstamm", f"E{i}", phone, "contact", "German mobile number")
        secret("Kundenstamm", f"F{i}", addr, "contact", "street + PLZ + city")

    # ---------------------------------------------------------------- 2. HR (DE, Art.9)
    ws2 = wb.create_sheet("Personal Vertraulich")
    ws2.append(["Mitarbeiter", "Abteilung", "Diagnose", "Konfession", "Gewerkschaft", "Vermerk"])
    for i in range(2, 8):
        fn, ln = rng.choice(de_people)
        diag, conf, uni = rng.choice(DIAGNOSES), rng.choice(CONFESSIONS), rng.choice(UNIONS)
        ws2.append([f"{fn} {ln}", rng.choice(DEPTS), diag, conf, uni,
                    f"Wiedereingliederung ab dem Folgequartal; Ruecksprache mit {ln} erforderlich."])
        secret("Personal Vertraulich", f"A{i}", f"{fn} {ln}", "people", "full name in a cell")
        secret("Personal Vertraulich", f"C{i}", diag, "special_category", "Art.9 health data (DE)")
        secret("Personal Vertraulich", f"D{i}", conf, "special_category", "Art.9 religion (DE)")
        secret("Personal Vertraulich", f"E{i}", uni, "special_category", "Art.9 union membership (DE)")

    # ---------------------------------------------------------------- 3. Projektportfolio (DE prose)
    ws3 = wb.create_sheet("Projektportfolio")
    ws3.append(["Projektname", "Eingesetztes Tool", "Abteilung", "Lizenzgeber", "Beschreibung", "Produktgruppe"])
    for i in range(2, 10):
        proj, tool, dept, vend = (rng.choice(PROJECTS), rng.choice(TOOLS),
                                  rng.choice(DEPTS), rng.choice(VENDORS))
        (fn1, ln1), (fn2, ln2) = rng.sample(de_people, 2)
        cost = rng.choice(COSTS)
        desc = (f"Das Projekt {proj} loest die Altsysteme im {dept} ab. "
                f"Fachliche Leitung: Herr Dr. {fn1} {ln1}, technische Umsetzung durch Frau {fn2} {ln2}. "
                f"Budget {cost}; {tool} wird von {vend} lizenziert.")
        ws3.append([proj, tool, dept, vend, desc, rng.choice(["Vorsorge", "Sparen", "Kredit"])])
        secret("Projektportfolio", f"A{i}", proj, "internal_topical", "project name via header")
        secret("Projektportfolio", f"B{i}", tool, "internal_topical", "internal tool via header")
        secret("Projektportfolio", f"D{i}", vend, "internal_topical", "licensor via header")
        secret("Projektportfolio", f"E{i}", f"{fn1} {ln1}", "people", "1st of TWO names in one sentence")
        secret("Projektportfolio", f"E{i}", f"{fn2} {ln2}", "people", "2nd of TWO names in one sentence")

    # ---------------------------------------------------------------- 4. Protokolle (DE minutes)
    ws4 = wb.create_sheet("Protokolle")
    ws4.append(["Datum", "Gremium", "Protokollnotiz"])
    for i in range(2, 9):
        (fn1, ln1), (fn2, ln2), (fn3, ln3) = rng.sample(de_people, 3)
        proj, tool = rng.choice(PROJECTS), rng.choice(TOOLS)
        note = rng.choice([
            f"Die Abstimmung zwischen {ln1}, {ln2} und {ln3} erfolgte ohne Gegenstimme. "
            f"{ln1}s Team uebernimmt die Migration nach {tool}.",
            f"Frau {fn1} {ln1} berichtet zum Stand von {proj}; Herr {fn2} {ln2} ergaenzt zur Kostenlage. "
            f"Beschluss: Fortfuehrung bis zum Folgequartal.",
            f"Dr. {fn1} {ln1} und {fn2} {ln2} eskalieren die Verzoegerung bei {proj}. "
            f"{ln3} pruegt die vertraglichen Optionen mit dem Lizenzgeber.",
        ])
        ws4.append([f"{rng.randint(1,28):02d}.{rng.randint(1,12):02d}.2025",
                    rng.choice(["Lenkungsausschuss", "Projektboard", "Vorstandssitzung"]), note])
        for who in ({ln1, ln2, ln3}):
            secret("Protokolle", f"C{i}", who, "people", "bare surname inside minutes prose")
        secret("Protokolle", f"C{i}", proj if proj in note else tool, "internal_topical",
               "project/tool named in prose (propagation)")

    # ---------------------------------------------------------------- 5. Kostenplanung (DE figures)
    ws5 = wb.create_sheet("Kostenplanung")
    ws5.append(["Projekt", "Verantwortlich", "Planwert", "Kommentar"])
    for i in range(2, 9):
        proj = rng.choice(PROJECTS)
        fn, ln = rng.choice(de_people)
        cost = rng.choice(COSTS)
        ws5.append([proj, f"{fn} {ln}", cost,
                    f"Hochrechnung {cost} bis Jahresende; Freigabe durch {ln} steht aus."])
        secret("Kostenplanung", f"B{i}", f"{fn} {ln}", "people", "owner name beside a cost figure")
        secret("Kostenplanung", f"A{i}", proj, "internal_topical", "project via header")
    # Decoys go in a NEUTRALLY-headed column, not in "Kommentar". A Kommentar/
    # Beschreibung column is a DESCRIPTION column by design -- the tool summarizes
    # it wholesale precisely because free-text columns hide confidential prose --
    # so a decoy placed there would be "failed" for behaviour that is correct.
    ws5["E1"] = "Status"
    ws5.append(["Gesamt", "", "TEUR 9.400", "", decoy(
        "Kostenplanung", "E9", "Die Projektlaufzeit betraegt 36 Monate.",
        "duration, contains no personal data")])
    ws5.append(["Reserve", "", "EUR 66.450", "", decoy(
        "Kostenplanung", "E10", "Ruecklage von 66450 Euro gebildet.",
        "5-digit amount, not a postal code")])
    ws5.append(["Kennzahl", "", "12%", "", decoy(
        "Kostenplanung", "E11", "Total cost of ownership was reduced by 12 percent.",
        "financial prose, no PII")])

    # ---------------------------------------------------------------- 6. Client Register UK (EN)
    ws6 = wb.create_sheet("Client Register UK")
    ws6.append(["Surname", "First name", "IBAN", "Nationality", "Religion",
                "Medical condition", "Trade union"])
    en_rows = [
        ("Okonkwo", "Grace", "Nigerian", "Roman Catholic", "type 2 diabetes, insulin dependent", "UNISON"),
        ("Fitzgerald", "Liam", "Irish", "Protestant", "chronic back pain", "Unite the Union"),
        ("Haddad", "Amina", "Lebanese", "Muslim", "diagnosed with multiple sclerosis", "Trades Union Congress"),
        ("Nakamura", "Daniel", "Japanese", "Buddhist", "recovering from a cardiac procedure", "UNISON"),
    ]
    for i, (ln, fn, nat, rel, med, uni) in enumerate(en_rows, start=2):
        ws6.append([ln, fn, iban(rng.randint(10**8, 10**9)), nat, rel, med, uni])
        secret("Client Register UK", f"A{i}", ln, "people", "bare surname, ENGLISH sheet")
        secret("Client Register UK", f"D{i}", nat, "special_category", "Art.9 ethnic origin (EN)")
        secret("Client Register UK", f"E{i}", rel, "special_category", "Art.9 religion (EN)")
        secret("Client Register UK", f"F{i}", med, "special_category", "Art.9 health data (EN)")
        secret("Client Register UK", f"G{i}", uni, "special_category", "Art.9 union membership (EN)")

    # ---------------------------------------------------------------- 7. Board Minutes EN (prose)
    ws7 = wb.create_sheet("Board Minutes EN")
    ws7.append(["Date", "Topic", "Note"])
    for i in range(2, 8):
        (fn1, ln1), (fn2, ln2) = rng.sample(en_people, 2)
        proj, vend, cost = rng.choice(PROJECTS), rng.choice(VENDORS), rng.choice(COSTS)
        # The note and the values planted for it are built TOGETHER: a template that
        # writes only surnames must not have full names recorded against it, or the
        # manifest measures recall against text that is not in the file.
        if rng.random() < 0.5:
            note = (f"Ms {fn1} {ln1} and Mr {fn2} {ln2} approved the {proj} business case at {cost}. "
                    f"The {vend} contract renews in the next quarter.")
            names = [f"{fn1} {ln1}", f"{fn2} {ln2}"]
            why = "full name, ENGLISH prose, two people in one sentence"
        else:
            note = (f"{ln1} raised a concern about the {proj} cost projection of {cost}; "
                    f"{ln2} will review the {vend} statement of work before sign-off.")
            names = [ln1, ln2]
            why = "bare surname, ENGLISH prose, two people in one sentence"
        ws7.append([f"2025-{rng.randint(1,12):02d}-{rng.randint(1,28):02d}",
                    rng.choice(["Investment", "Vendor review", "Risk"]), note])
        for nm in names:
            secret("Board Minutes EN", f"C{i}", nm, "people", why)
        secret("Board Minutes EN", f"C{i}", proj, "internal_topical", "project named in ENGLISH prose")

    # ---------------------------------------------------------------- 8. Vendor Contracts (EN topical)
    ws8 = wb.create_sheet("Vendor Contracts")
    ws8.append(["Supplier", "Project", "Department", "Description"])
    for i in range(2, 7):
        vend, proj = rng.choice(VENDORS), rng.choice(PROJECTS)
        ws8.append([vend, proj, rng.choice(["Group Risk", "Treasury", "Compliance"]),
                    f"Quarterly review of the {proj} migration backlog and the renewal options."])
        secret("Vendor Contracts", f"A{i}", vend, "internal_topical", "supplier via ENGLISH header")
        secret("Vendor Contracts", f"B{i}", proj, "internal_topical", "project via ENGLISH header")

    # ---------------------------------------------------------------- 9. Strategie (decoy-heavy)
    ws9 = wb.create_sheet("Strategie Notizen")
    ws9.append(["Thema", "Entscheidung"])
    s_fn, s_ln = de_people[0]
    s_tool = TOOLS[0]
    ws9.append(["Zielbild 2027",
                f"Ausstieg aus dem Segment Kapitalmarkt bis Q3. Verantwortlich: {s_fn} {s_ln}."])
    secret("Strategie Notizen", "B2", f"{s_fn} {s_ln}", "people", "name in free text (propagation)")
    ws9.append(["Toolstrategie",
                f"{s_tool} wird konzernweit abgeloest; die Migration beginnt im Folgejahr."])
    secret("Strategie Notizen", "B3", s_tool, "internal_topical", "tool in free text (propagation)")
    for label, text, why in [
        ("Marktumfeld", "The Great Depression started in 1929 and reshaped banking.",
         "historical prose, not health data"),
        ("Gegenpartei", "Credit Union: Nationwide Building Society", "counterparty, not union membership"),
        ("Methodik", "An orthodox approach to risk management was applied.",
         "ordinary adjective, not religion"),
        ("Technik", "Race conditions in the settlement engine were fixed.",
         "engineering term, not ethnic origin"),
        ("Regulatorik", "European Union regulations apply to this transaction.",
         "not union membership"),
        ("Effizienz", "Die Effizienz der Reaktionszeiten war zufriedenstellend.",
         "German nominalizations, not names"),
        ("Vertrag", "Im Anhang 2 sind die Konditionen beschrieben.",
         "prepositional phrase, not a street address"),
        ("Referenz", "Vertragsnummer VN-2024-0087651234 wurde vergeben.",
         "internal contract number, not an IBAN"),
        ("Termin", "Zum Stichtag 31.12.2025 betrug der Saldo null.",
         "date anchor, not a street address"),
    ]:
        r = ws9.max_row + 1
        ws9.append([label, decoy("Strategie Notizen", f"B{r}", text, why)])

    # ---------------------------------------------------------------- 10. hard surfaces
    ws10 = wb.create_sheet("Abrechnung")
    ws10.append(["Position", "Hinweis"])
    f_fn, f_ln = de_people[1]
    ws10["B2"] = f'="Sachbearbeiter: "&"{f_fn} {f_ln}"'
    secret("Abrechnung", "B2", f"{f_fn} {f_ln}", "people", "name inside a FORMULA literal")
    comment_iban = iban(rng.randint(10**8, 10**9))
    ws10["A2"] = "Sammelbuchung"
    ws10["A2"].comment = Comment(f"Rueckfrage zur IBAN {comment_iban}", "Revision")
    secret("Abrechnung", "A2", comment_iban, "financial_ids", "IBAN inside a cell COMMENT")

    h_fn, h_ln = de_people[2]
    hidden_name = f"Archiv {h_fn} {h_ln}"[:31]
    ws11 = wb.create_sheet(hidden_name)
    ws11.sheet_state = "hidden"
    ws11.append(["Vermerk"])
    ws11.append([f"Altbestand {h_fn} {h_ln}, Auskunft nur nach Ruecksprache mit {h_ln}."])
    secret(hidden_name, "A2", f"{h_fn} {h_ln}", "people", "name on a HIDDEN sheet")
    must_catch.append({"sheet": hidden_name, "cell": "<sheet title>", "value": f"{h_fn} {h_ln}",
                       "data_class": "people", "why": "sensitive SHEET NAME (xl/workbook.xml)"})

    # ------------------------------------------------- 12. DB_0_Metadaten (people + links)
    # English-headed people columns and internal links -- the two recall gaps the
    # 2026-07-27 audit found. Owner/Einreicher/MDX_Lead/MDX_Proxy hold people; the
    # three link columns hold a hostname plus a document pointer.
    ws12 = wb.create_sheet("DB_0_Metadaten")
    ws12.append(["Projekt_ID", "Titel", "Geschaeftsbereich", "Abteilung", "Einreicher", "Owner",
                 "MDX_Lead", "MDX_Proxy", "Confluence_Link", "Prozess_URL", "Aktuelle_Phase"])
    for i, title in enumerate(INNO_TITLES, start=2):
        (of, ol), (ef, el_) = rng.sample(de_people, 2)
        lead, proxy = rng.sample(LEADS, 2)
        owner, einreicher = f"{of} {ol}", f"{ef} {el_}"
        host = rng.choice(INNO_HOSTS)
        conf = f"https://{host}/x/{rng.randint(10**6, 10**7)}"
        purl = f"https://{rng.choice(INNO_HOSTS)}/apps/workflow.nsf/doc/{rng.randint(10**7, 10**8)}?open"
        ws12.append([f"INNO-26-{i:03d}", title, rng.choice(DIVISIONS), rng.choice(DEPTS),
                     einreicher, owner, lead, proxy, conf, purl, rng.choice(RELEVANZ_LIST)])
        secret("DB_0_Metadaten", f"E{i}", einreicher, "people", "full name, ENGLISH-headed people column")
        secret("DB_0_Metadaten", f"F{i}", owner, "people", "full name under 'Owner'")
        secret("DB_0_Metadaten", f"G{i}", lead, "people", "BARE FIRST NAME alone in a cell")
        secret("DB_0_Metadaten", f"H{i}", proxy, "people", "BARE FIRST NAME alone in a cell")
        secret("DB_0_Metadaten", f"I{i}", conf, "bank_internal", "internal link: host + document id")
        secret("DB_0_Metadaten", f"J{i}", purl, "bank_internal", "internal deep link with query string")

    # ------------------------------------------------- 13. DB_Log (one name, many rows)
    # An audit log: the SAME person in a `User` column on every row. On the reported
    # workbook one name sat in such a column 318 times and left in the clear.
    ws13 = wb.create_sheet("DB_Log")
    ws13.append(["Zeitstempel", "Projekt_ID", "User", "Feld_Name", "Alter_Wert", "Neuer_Wert"])
    log_fn, log_ln = de_people[4]
    log_user = f"{log_fn} {log_ln}"
    for i in range(2, 22):
        ws13.append([f"2026-0{rng.randint(1,9)}-{rng.randint(10,28)} {rng.randint(8,17)}:{rng.randint(10,59)}",
                     f"INNO-26-{rng.randint(2,9):03d}", log_user,
                     decoy("DB_Log", f"D{i}", f"Beschreibung_{i}", "snake_case schema field id, not a name"),
                     rng.choice(STATUS_LIST), rng.choice(STATUS_LIST)])
        secret("DB_Log", f"C{i}", log_user, "people", "same person repeated down a 'User' column")

    # ------------------------------------------------- 14. DB_Setup (controlled vocabulary)
    # The dropdown source lists, declared as REAL Excel data validations so the
    # workbook's own controlled vocabulary is machine-readable. Every value is a
    # decoy: a scale label or a status is never personal data, and these dominated
    # the measured false positives (9 such values were 79% of the noise).
    ws14 = wb.create_sheet("DB_Setup")
    ws14.append(["Liste_3Pkt_Skala", "Liste_Standard_Status", "Liste_Relevanz", "Persona_Rollen"])
    for i, row in enumerate(
        zip(SCALE_3PKT, STATUS_LIST, RELEVANZ_LIST + [""], PERSONA_ROLES + [""]), start=2
    ):
        ws14.append(list(row))
        for col, val in zip("ABCD", row):
            if val:
                decoy("DB_Setup", f"{col}{i}", val, "controlled-vocabulary dropdown value, never a name")

    # ------------------------------------------------- 15. DB_2_Concept (enum columns + jargon)
    # Where the vocabulary is USED: the same handful of labels repeated on every row,
    # which is what makes a repetition signal possible. `Einschaetzung` is
    # deliberately a NEUTRAL header -- not a Kommentar/Beschreibung/Notiz column --
    # so a decoy there is a true false positive rather than the wholesale
    # summarization a description column is supposed to get (see the note at
    # Kostenplanung above).
    ws15 = wb.create_sheet("DB_2_Concept")
    ws15.append(["Projekt_ID", "GB_Fit_Marke", "GB_Fit_Innovation", "Ownership_geklaert_Status",
                 "Einschaetzung", "Zielmetrik"])
    for i in range(2, 14):
        marke, inno = rng.choice(SCALE_3PKT), rng.choice(SCALE_3PKT)
        own = rng.choice(STATUS_LIST)
        jargon = JARGON_DE[(i - 2) % len(JARGON_DE)]
        ws15.append([f"INNO-26-{rng.randint(2,9):03d}", marke, inno, own,
                     f"{jargon}: die Umsetzung erfolgt schrittweise.", f"Zielwert_{i}"])
        decoy("DB_2_Concept", f"B{i}", marke, "3-point scale label repeated down a column")
        decoy("DB_2_Concept", f"D{i}", own, "status value under an 'Ownership...' header")
        decoy("DB_2_Concept", f"E{i}", jargon, "German compound noun / business jargon, not a person")
    # Real data validations, pointing at the DB_Setup lists. This is the signal a
    # precision pass can read INSTEAD of guessing: the workbook declares, in its own
    # XML, that these columns may only hold those values.
    for col, ref in (("B", "$A$2:$A$5"), ("C", "$A$2:$A$5"), ("D", "$B$2:$B$7")):
        dv = DataValidation(type="list", formula1=f"DB_Setup!{ref}", allow_blank=True)
        ws15.add_data_validation(dv)
        dv.add(f"{col}2:{col}200")

    # ------------------------------------------------- 16. DB_1_Ideation (prose + multi-value)
    # Long German free text under a genuine Beschreibung column (so the whole cell is
    # a description by design), with names planted inside it, plus a MULTI-VALUE cell
    # written the way Excel writes one: parts joined by the U+001E record separator,
    # which openpyxl hands back as a literal `_x001E_` escape.
    ws16 = wb.create_sheet("DB_1_Ideation")
    ws16.append(["Projekt_ID", "Beschreibung", "Kern_Problem_Pain", "Persona_1_Rolle"])
    for i in range(2, 8):
        (f1, l1), (f2, l2) = rng.sample(de_people, 2)
        tool = rng.choice(TOOLS)
        ws16.append([
            f"INNO-26-{rng.randint(2,9):03d}",
            f"Fachliche Begleitung durch {f1} {l1}; technische Bewertung durch Frau {f2} {l2}. "
            f"Die Anbindung an {tool} ist Voraussetzung fuer den Piloten.",
            f"Manuelle Nacharbeit bindet Kapazitaet; {l1} schaetzt den Folgeaufwand als hoch ein.",
            rng.choice(PERSONA_ROLES),
        ])
        secret("DB_1_Ideation", f"B{i}", f"{f1} {l1}", "people", "name inside a description cell")
        secret("DB_1_Ideation", f"B{i}", f"{f2} {l2}", "people", "second name in the same description")
        secret("DB_1_Ideation", f"C{i}", l1, "people", "bare surname in a second free-text column")
    # A multi-value cell whose parts are EMPTY: nothing but separators. Excel writes
    # the separator as a literal `_x001E_` escape, whose x and E are word characters,
    # so an emptiness guard written as `^[\W\d_]*$` cannot match it -- measured, 7
    # entirely empty cells were flagged at the auto-accept tier because of this.
    sep = "_x001E_"
    r = ws16.max_row + 1
    ws16.cell(row=r, column=1, value="INNO-26-099")
    ws16.cell(row=r, column=2, value=sep * 4)
    decoy("DB_1_Ideation", f"B{r}", sep * 4, "multi-value cell with only EMPTY parts -- no content at all")
    # ...and one whose parts are real, with a name in the second part: the same escape
    # previously fused onto the adjacent token and got the whole span rejected as a
    # snake_case identifier, hiding the name entirely.
    mv_fn, mv_ln = de_people[5]
    r += 1
    ws16.cell(row=r, column=1, value="INNO-26-098")
    ws16.cell(row=r, column=2, value=f"Kein Beitrag{sep}Fachkontakt {mv_fn} {mv_ln}{sep}offen")
    secret("DB_1_Ideation", f"B{r}", f"{mv_fn} {mv_ln}", "people",
           "name fused to an Excel _x001E_ escape in a multi-value cell")

    # De-duplicate: the same value planted twice at the same cell adds nothing.
    seen = set()
    deduped = []
    for s in must_catch:
        k = (s["sheet"], s["cell"], s["value"])
        if k not in seen:
            seen.add(k)
            deduped.append(s)
    must_catch = deduped

    xlsx_path = out_dir / "audit_workbook.xlsx"
    wb.save(xlsx_path)

    manifest = {
        "seed": seed,
        "workbook": xlsx_path.name,
        "sheets": wb.sheetnames,
        "must_catch": must_catch,
        "must_not_catch": must_not_catch,
    }
    (out_dir / "audit_manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    print(f"workbook : {xlsx_path}  ({xlsx_path.stat().st_size/1024:.0f} KB)")
    print(f"sheets   : {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        n = sum(1 for m in must_catch if m["sheet"] == s)
        print(f"             {s:24} {n:3} planted")
    print(f"planted  : {len(must_catch)} secrets, {len(must_not_catch)} decoys")
    return 0


if __name__ == "__main__":
    target = Path(sys.argv[1] if len(sys.argv) > 1 else "tests/fixtures/generated").resolve()
    raise SystemExit(main(target, int(sys.argv[2]) if len(sys.argv) > 2 else SEED))
