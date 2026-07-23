"""Run-level find/replace shared by docx and pptx handlers (and the raw-XML
adapter for footnotes/comments). Works on any object with a get/set `.text`.

Also owns the OOXML AUXILIARY SURFACES (bottom of this file) -- reader-visible
text that no format library exposes through its document object model, and which
was therefore invisible to the scan AND to the handler-based output re-scan."""

from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path

import openpyxl
from lxml import etree

from .. import xmlsafe
from ..actions import decisions_lookup, resolve_replacement
from ..core import detect_unit
from ..models import ProcessingError, TextUnit


def runs_text_and_spans(runs: list) -> tuple[str, list[tuple[object, int, int]]]:
    full_text = ""
    spans = []
    pos = 0
    for run in runs:
        t = run.text or ""
        spans.append((run, pos, pos + len(t)))
        full_text += t
        pos += len(t)
    return full_text, spans


def apply_span_replacement(spans: list[tuple[object, int, int]], start: int, end: int, replacement: str) -> None:
    overlapping = [(r, rs, re) for (r, rs, re) in spans if re > start and rs < end]
    if not overlapping:
        return
    first = True
    for run, rs, re in overlapping:
        local_start = max(start, rs) - rs
        local_end = min(end, re) - rs
        text = run.text or ""
        if first:
            run.text = text[:local_start] + replacement + text[local_end:]
            first = False
        else:
            run.text = text[:local_start] + text[local_end:]


def apply_findings_to_runs(runs: list, findings: list, decisions: dict, mapping_store) -> None:
    """Applies findings to runs in descending start order so earlier offsets stay valid."""
    if not findings:
        return
    _, spans = runs_text_and_spans(runs)
    for f in sorted(findings, key=lambda f: -f.start):
        action = decisions_lookup(decisions, f.entity_type, f.value)
        replacement = resolve_replacement(f.entity_type, f.value, action, mapping_store)
        if replacement is None:
            continue
        apply_span_replacement(spans, f.start, f.end, replacement)


class XmlRunAdapter:
    """Adapts a single OOXML `<w:t>`/`<a:t>` element to the run-like `.text` interface."""

    def __init__(self, t_element):
        self._t = t_element

    @property
    def text(self) -> str:
        return self._t.text or ""

    @text.setter
    def text(self, value: str) -> None:
        self._t.text = value


def redact_text(text: str, analyzer, config, decisions: dict, mapping_store) -> str:
    """Splice-redacts a standalone string (a hyperlink target, an embedded-workbook
    cell). detect_unit returns an overlap-resolved set, so right-to-left splicing
    is safe."""
    result = text
    for f in sorted(detect_unit(analyzer, TextUnit(id="tmp", text=text), config), key=lambda f: -f.start):
        action = decisions_lookup(decisions, f.entity_type, f.value)
        replacement = resolve_replacement(f.entity_type, f.value, action, mapping_store)
        if replacement is None:
            continue
        result = result[: f.start] + replacement + result[f.end :]
    return result


# --- OOXML auxiliary surfaces ------------------------------------------------
#
# THE FALSE-CLEAN HOLE these close: verify_output re-scans the written output
# through the SAME handler that produced it, so any text location the handler is
# blind to on the way IN is equally invisible on the way OUT -- and the literal
# backstop only re-checks values that were already DECIDED, which PII the scan
# never surfaced can never be. A value living only in one of the surfaces below
# was therefore reported "verified clean" without ever being looked at:
#
#   * an EXTERNAL hyperlink TARGET -- only the display text was ever scanned, but
#     the URL is where "mailto:hans.mueller@bank.de" actually lives, and it is a
#     relationship ATTRIBUTE that no itertext() sweep reads;
#   * CHART parts (title, cached category/series labels) -- python-docx/-pptx walk
#     shapes and paragraphs, never `charts/chart1.xml`;
#   * the EMBEDDED WORKBOOK behind a chart -- a whole OOXML package nested inside
#     a compressed part, so even a raw byte scan of the outer file cannot see it.
#
# Handled generically on the raw package so the pass is HANDLER-INDEPENDENT and
# identical for docx / xlsx / pptx.

_A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_C_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
_A_T = f"{{{_A_NS}}}t"
_A_P = f"{{{_A_NS}}}p"
_C_V = f"{{{_C_NS}}}v"

_CHART_PART_RE = re.compile(r"(?:^|/)charts/chart[^/]*\.xml$", re.IGNORECASE)
_EMBEDDED_WORKBOOK_RE = re.compile(r"(?:^|/)embeddings/[^/]+\.xls[xm]$", re.IGNORECASE)
# A c:v holding a plain number is the chart's NUMBER CACHE. Splicing a token into
# it would corrupt the plot, and a bare figure carries no identity on its own --
# identifying digit strings (account numbers) live in the string caches, which are
# scanned. Only non-numeric cached values are treated as text.
_NUMERIC_RE = re.compile(r"^[+-]?\d+(?:[.,]\d+)?(?:[eE][+-]?\d+)?$")


def _parse(blob: bytes):
    try:
        return xmlsafe.fromstring(blob)
    except etree.XMLSyntaxError:
        return None


def _external_rels(tree):
    """Relationship elements whose target is EXTERNAL -- a hyperlink URL, a linked
    file path. Internal targets are package paths and never carry PII."""
    for el in tree.iter():
        if not isinstance(el.tag, str) or not el.tag.endswith("}Relationship"):
            continue
        if el.get("TargetMode") == "External" and (el.get("Target") or "").strip():
            yield el


# A URI's scheme (and authority marker), e.g. "mailto:", "https://", "file:///".
_URI_PREFIX_RE = re.compile(r"^[A-Za-z][A-Za-z0-9+.\-]*:(?://+)?")


def _safe_relationship_target(original: str, redacted: str) -> str:
    """The redacted Target, kept a VALID URI.

    Two ways the naive rewrite produced a broken one. A Target that is entirely PII
    redacts to a bare token, dropping the scheme, so the original scheme is put
    back. And the token's square brackets are reserved by RFC 3986 for IPv6
    literals, so they are percent-encoded. Word treats an external relationship
    with an invalid Target as unreadable content and offers to "repair" the file --
    which for this tool means the colleague receives what looks like a corrupt
    document, so this is worth the ugliness."""
    if redacted == original:
        return redacted
    prefix = _URI_PREFIX_RE.match(original)
    if prefix and not _URI_PREFIX_RE.match(redacted):
        redacted = prefix.group(0) + redacted
    return redacted.replace("[", "%5B").replace("]", "%5D")


def _chart_text_groups(tree):
    """Groups of text-bearing chart elements. Each `a:p` paragraph's runs form ONE
    group, so a title split across formatting runs rejoins before detection (a name
    split mid-word would otherwise be undetectable); every other text node stands
    alone."""
    grouped: set[int] = set()
    for para in tree.iter(_A_P):
        group = list(para.iter(_A_T))
        if group:
            grouped.update(id(t) for t in group)
            yield group
    for t in tree.iter(_A_T):
        if id(t) not in grouped:
            yield [t]
    for v in tree.iter(_C_V):
        text = (v.text or "").strip()
        if text and not _NUMERIC_RE.match(text):
            yield [v]


def _open_package(path: Path) -> zipfile.ZipFile:
    """The document's zip container, or a loud failure. Input documents are
    UNTRUSTED, so "cannot open it" is precisely the case that must not pass
    quietly: swallowing it meant every auxiliary surface (hyperlink targets, chart
    text, embedded workbooks) went unscanned AND unredacted while the run reported
    success."""
    try:
        return zipfile.ZipFile(path)
    except (zipfile.BadZipFile, OSError) as exc:
        raise ProcessingError(
            f"'{path.name}' could not be opened as an Office package ({exc}). It cannot be "
            "checked for hidden text (hyperlink targets, charts, embedded workbooks), so no "
            "output was written."
        ) from exc


def _load_embedded_workbook(data: bytes):
    """The embedded workbook, or a loud failure. An embedding we cannot read is an
    embedding we cannot scan or redact -- and a chart's source workbook is a normal
    place for a customer list to live."""
    try:
        return openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    except Exception as exc:  # noqa: BLE001
        raise ProcessingError(
            f"An embedded workbook inside this document could not be read ({exc}), so its "
            "contents could be neither checked nor redacted. No output was written."
        ) from exc


def _embedded_workbook_strings(data: bytes):
    """(sheet, coordinate, text) for every string cell of an embedded workbook."""
    wb = _load_embedded_workbook(data)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip():
                        yield ws.title, cell.coordinate, cell.value
    finally:
        wb.close()


def aux_text_units(path: Path) -> list[TextUnit]:
    """TextUnits for the auxiliary surfaces above, so their PII is surfaced to the
    reviewer instead of silently riding along. Fails loud on anything it cannot
    read: the backstop only re-checks values that were already DECIDED, so a
    surface we never opened is a surface nothing checks."""
    units: list[TextUnit] = []
    with _open_package(path) as zf:
        for name in zf.namelist():
            if name.endswith(".rels"):
                tree = _parse(zf.read(name))
                if tree is None:
                    continue
                for i, rel in enumerate(_external_rels(tree)):
                    units.append(TextUnit(id=f"aux|{name}|link{i}", text=rel.get("Target")))
            elif _CHART_PART_RE.search(name):
                tree = _parse(zf.read(name))
                if tree is None:
                    continue
                for i, group in enumerate(_chart_text_groups(tree)):
                    text = "".join(el.text or "" for el in group)
                    if text.strip():
                        units.append(TextUnit(id=f"aux|{name}|{i}", text=text))
            elif _EMBEDDED_WORKBOOK_RE.search(name):
                for sheet, coord, text in _embedded_workbook_strings(zf.read(name)):
                    units.append(TextUnit(id=f"aux|{name}|{sheet}!{coord}", text=text))
    return units


def _redact_embedded_workbook(data: bytes, analyzer, config, decisions: dict, mapping_store) -> bytes | None:
    """Redacted bytes for an embedded workbook, or None if unchanged. An embedding
    that cannot be read fails loud (see `_load_embedded_workbook`)."""
    wb = _load_embedded_workbook(data)
    changed = False
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip():
                        new = redact_text(cell.value, analyzer, config, decisions, mapping_store)
                        if new != cell.value:
                            cell.value = new
                            changed = True
        if not changed:
            return None
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        wb.close()


def apply_aux_parts(path: Path, analyzer, config, decisions: dict, mapping_store) -> None:
    """Redacts the auxiliary surfaces of an already-written OOXML output, in place.
    Runs AFTER the format library has saved the file, because these parts survive
    that save untouched -- which is exactly why they leaked. Fails loud rather than
    returning quietly on a package it cannot open: silently skipping the redaction
    of a surface would ship the very PII this pass exists to remove."""
    with _open_package(path) as zf:
        names = zf.namelist()
        contents = {n: zf.read(n) for n in names}

    changed = False
    for name in names:
        if name.endswith(".rels"):
            tree = _parse(contents[name])
            if tree is None:
                continue
            part_changed = False
            for rel in _external_rels(tree):
                target = rel.get("Target")
                new = _safe_relationship_target(
                    target, redact_text(target, analyzer, config, decisions, mapping_store)
                )
                if new != target:
                    rel.set("Target", new)
                    part_changed = True
            if part_changed:
                contents[name] = etree.tostring(tree, xml_declaration=True, encoding="UTF-8", standalone=True)
                changed = True
        elif _CHART_PART_RE.search(name):
            tree = _parse(contents[name])
            if tree is None:
                continue
            part_changed = False
            for group in _chart_text_groups(tree):
                runs = [XmlRunAdapter(el) for el in group]
                text = runs_text_and_spans(runs)[0]
                if not text.strip():
                    continue
                findings = detect_unit(analyzer, TextUnit(id="tmp", text=text), config)
                if not findings:
                    continue
                apply_findings_to_runs(runs, findings, decisions, mapping_store)
                part_changed = True
            if part_changed:
                contents[name] = etree.tostring(tree, xml_declaration=True, encoding="UTF-8", standalone=True)
                changed = True
        elif _EMBEDDED_WORKBOOK_RE.search(name):
            new_blob = _redact_embedded_workbook(contents[name], analyzer, config, decisions, mapping_store)
            if new_blob is not None:
                contents[name] = new_blob
                changed = True

    if changed:
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as out:
            for name in names:
                out.writestr(name, contents[name])
