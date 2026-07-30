from __future__ import annotations

import functools
import re
from dataclasses import replace
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from .. import language, taxonomy
from ..actions import decisions_lookup, resolve_replacement, token_label
from ..core import (
    _CORROBORATION_ONLY_ENTITIES,
    _GATED_NER_SOURCES,
    _NER_ENTITIES,
    _resolve_overlaps,
    detect_unit,
    is_given_name,
    neutralize_structural_noise,
    precompute_nlp_artifacts,
)
from ..engine import DEFAULT_LANGUAGES
from ..gliner_recognizer import prime_gliner
from ..models import CellInfo, ColumnInfo, Finding, ProcessingError, TextUnit
from .run_replace import apply_aux_parts, aux_text_units

EXTENSIONS = (".xlsx", ".xlsm", ".xls")

# A column header that declares its cells are people. This is the one place a
# bare surname legitimately appears with no prose around it, and it is exactly
# where NER collapses: measured, de_core_news_lg finds only ~35% of ordinary
# German surnames (Müller, Weber, Bauer) in a bare cell, because its training
# gives it nothing to lean on without a sentence.
#
# The header is stronger evidence than any model: the spreadsheet's own author
# labelled the column. Note this does NOT work via Presidio's context boost --
# that only lifts PATTERN recognizers, and spaCy NER gets no boost from it at
# all, which is why "Kunde: Müller" in a cell still missed.
# Built-in column-header stems that declare a column holds PEOPLE. Matched
# case-insensitively as a SUBSTRING, so "leiter" covers Projekt-/Team-/Abteilungs-
# leiter and "berechtigt" covers zeichnungs-/bevollmächtigt forms. The shipped set
# was too narrow for a real "database" workbook (it missed Projektleiter, Betreuer,
# Verantwortlich, ...), so a name column with such a header leaked ~65% via NER
# alone. Extend per workbook via config["name_column_headers"] (Settings > Detection).
_NAME_HEADER_TERMS = (
    "name", "kunde", "kundin", "inhaber", "empfänger", "empfaenger",
    "sachbearbeiter", "ansprechpartner", "berater", "beraterin", "mitarbeiter",
    "antragsteller", "vertragspartner", "begünstigter", "beguenstigter",
    # widened: common German business / bank name-column headers.
    "projektleiter", "leiter", "betreuer", "verantwortlich", "referent",
    "gesellschafter", "geschäftsführer", "geschaeftsfuehrer", "prokurist",
    "zeichnungsberechtigt", "bevollmächtigt", "bevollmaechtigt", "berechtigt",
    "unterzeichner", "auftraggeber", "eigentümer", "eigentuemer",
    "vorname", "nachname", "familienname", "teilnehmer", "kontaktperson",
)

# WHOLE-WORD people-column stems -- a SECOND class, not more entries above.
#
# Substring matching is deliberate and load-bearing for the German stems (one
# "leiter" covers Projekt-/Team-/Abteilungsleiter), but it is wrong for these:
# measured on a real workbook, a substring "owner" also matches
# "Ownership_geklaert_Status", whose values are statuses ("Ausstehend"), and
# spaCy tags those ADJ -- which _NON_NAME_POS does not reject (it must not: Klein,
# Gross, Lang and Weiss are all real surnames spaCy can tag ADJ). So the shape
# gate cannot save us here and the header match itself has to be precise.
#
# Matched with a lowercase-letter boundary on each side rather than \b, because
# the real headers are `_`-joined ("Rollout_Owner", "MDX_Lead") and `_` is a word
# character -- \b would reject exactly the shape a database sheet uses.
#
# These are ENGLISH-language equivalents of the German stems above plus the four
# words this workbook actually uses. English headers were the whole recall gap:
# every name under Owner/Einreicher/MDX_Lead/MDX_Proxy leaked, because the
# whole-cell override never fired and bare spaCy misses a first name in a cell.
# Measured on that workbook: 83 real people recovered, 0 false positives.
#
# "user"/"users" earns its place the same way: an audit-log sheet carried the
# person who made each change in a column headed `User`, and on the reported
# workbook ONE name sat there 318 times. It was caught only because spaCy happened
# to recognise it, which is exactly the fragile path this override exists to
# replace. A login id rather than a name in such a column is harmless -- it is
# lowercase, so _looks_like_name rejects it.
_NAME_HEADER_WORDS = (
    "owner", "einreicher", "lead", "leads", "proxy", "submitter", "requester",
    "assignee", "reporter", "approver", "author", "creator", "manager",
    "contact", "kontakt", "responsible", "holder", "beneficiary", "signatory",
    "employee", "participant", "bearbeiter", "sponsor", "recipient", "applicant",
    "user", "users", "bearbeitet", "geaendert",
)


@functools.lru_cache(maxsize=8)
def _name_header_re(extra_terms: tuple[str, ...] = ()):
    """Compiled people-column-header matcher for the built-in stems plus any
    workbook-specific extras. lru_cached on the extra-terms tuple so the per-cell
    hot path never recompiles.

    User-supplied `extra_terms` keep the documented SUBSTRING semantics (see the
    `name_column_headers` comment in default_recognizers.yaml) -- only the
    built-in _NAME_HEADER_WORDS class is boundary-matched."""
    loose = _NAME_HEADER_TERMS + tuple(t.strip().lower() for t in extra_terms if t.strip())
    words = "|".join(re.escape(t) for t in _NAME_HEADER_WORDS)
    alts = [re.escape(t) for t in loose] + [rf"(?<![a-z])(?:{words})(?![a-z])"]
    return re.compile("|".join(alts), re.IGNORECASE)
# Cell contents that a name column can still hold but which are not names.
_NOT_A_NAME = re.compile(r"^[\W\d_]*$|^(unbekannt|n/?a|keine?|leer|-{1,3}|divers)$", re.IGNORECASE)
_NAME_COLUMN_SCORE = 0.8

# A value that is ONE snake_case token is a field/placeholder identifier, not a
# real name: measured, a "Team" column held team_1 .. team_5 and each was claimed
# as a DEPARTMENT at the auto-accept tier. The whole-cell PERSON override already
# rejects these via _looks_like_name; the TOPICAL override had no shape gate at all.
#
# Deliberately "is a single snake_case token", NOT "contains an underscore": a real
# DESCRIPTION cell is prose that may well mention a snake_case field name, and
# refusing to claim it for that reason would leave a confidential description
# unsummarized -- trading a cosmetic false positive for an actual leak.
_SNAKE_TOKEN = re.compile(r"^\w*_\w[\w-]*$")


def _is_placeholder_token(value: str) -> bool:
    v = value.strip()
    return " " not in v and bool(_SNAKE_TOKEN.match(v))

# --- topical (non-personal) category detection --------------------------------
# Header-confirmed, so scored into the auto-accept tier; source-tagged so it is
# never mistaken for a bare NER guess (bypasses corroboration-only / noise
# filters, which only gate NER entity types).
_TOPICAL_SCORE = 0.9
# A gazetteer term must be name-shaped (short), not a whole prose description --
# long PROJECT-description cells are handled whole-cell, not propagated.
_MAX_GAZETTEER_LEN = 40


def _topical_categories(config: dict) -> dict:
    """{CATEGORY: {header_terms, terms}} from config, or {} when topical
    detection is disabled/absent."""
    t = config.get("topical") or {}
    if not t.get("enabled", True):
        return {}
    return t.get("categories") or {}


@functools.lru_cache(maxsize=8)
def _topical_header_res(categories_key: tuple):
    """Per-category compiled header matchers. lru_cached on a hashable
    ((CATEGORY, (term, ...)), ...) key so the per-cell hot path never recompiles.

    WORD-BOUNDARY matching (not bare substring, which the people-column matcher
    uses): a category column drives WHOLE-COLUMN redaction, so a false header
    match is high-impact. Substring matched 'gruppe' inside 'Produktgruppe'
    (a product group, not a department); \\b requires the term to stand as a word,
    so 'Gruppe'/'Team'/'Abteilung' match but 'Produktgruppe'/'Anwendungsfall' do
    not. German compounds where the category word is a suffix (Fachabteilung) are
    intentionally NOT matched -- add such headers explicitly if needed (favouring
    precision, since a wrong category column redacts every cell in it)."""
    out = {}
    for cat, terms in categories_key:
        cleaned = [t.strip().lower() for t in terms if t and t.strip()]
        if cleaned:
            out[cat] = re.compile(r"\b(?:" + "|".join(re.escape(t) for t in cleaned) + r")\b", re.IGNORECASE)
    return out


def _category_for_header(header: str | None, config: dict) -> str | None:
    """The topical category a column header declares, or None. First match wins
    in config order (deterministic)."""
    if not header:
        return None
    cats = _topical_categories(config)
    key = tuple((cat, tuple(spec.get("header_terms", []))) for cat, spec in cats.items())
    for cat, rx in _topical_header_res(key).items():
        if rx.search(header):
            return cat
    return None


def topical_gazetteer(path: Path, config: dict) -> list[tuple[str, str]]:
    """Auto-learn topical terms from the document's own structure: every
    name-shaped value in a column whose header maps to a category becomes a
    (category, value) the caller propagates document-wide. Derived identically
    at scan and apply (both call this), so scan/apply parity holds. Long prose
    (PROJECT descriptions) is excluded -- it is handled whole-cell, not
    propagated."""
    cats = _topical_categories(config)
    if not cats:
        return []
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    pairs: set[tuple[str, str]] = set()
    try:
        for ws in wb.worksheets:
            col_cat: dict[str, str] = {}
            for cell in next(ws.iter_rows(min_row=1, max_row=1), []):
                if isinstance(cell.value, str) and cell.value.strip():
                    cat = _category_for_header(cell.value.strip(), config)
                    if cat:
                        col_cat[get_column_letter(cell.column)] = cat
            if not col_cat:
                continue
            # Only NAME categories seed the gazetteer; DESCRIPTION (free text) is
            # whole-cell summarized, never propagated.
            col_cat = {col: cat for col, cat in col_cat.items() if cat in taxonomy.PROPAGATING_TOPICAL_TYPES}
            if not col_cat:
                continue
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.value in (None, ""):
                        continue  # read_only fills gaps with EMPTY_CELL (no .column)
                    cat = col_cat.get(get_column_letter(cell.column))
                    if not cat:
                        continue
                    v = str(cell.value).strip()
                    # A single snake_case token is a field/placeholder identifier,
                    # never a real team or tool NAME -- measured, a "Team" column
                    # held team_1 .. team_5 and each was learned as a DEPARTMENT and
                    # then propagated document-wide. Excluded here as well as in the
                    # whole-cell override, because the gazetteer is the path that
                    # actually produced them (it spreads a learned term everywhere,
                    # so one bad entry costs far more than one bad cell).
                    if _is_placeholder_token(v):
                        continue
                    if 2 <= len(v) <= _MAX_GAZETTEER_LEN and any(ch.isalpha() for ch in v):
                        pairs.add((cat, v))
    finally:
        wb.close()
    return sorted(pairs)

# German name particles that don't themselves need to be capitalized ("Klaus
# von Bergen", "Anna de Wit").
# Kept aligned with engine._NAME_PARTICLE, which the anchored patterns use: a
# particle this gate rejects but the anchors accept produces a cell the tool
# detects in prose and refuses to claim in a column ("van den Broek", "de la Cruz").
_NAME_PARTICLES = {
    "von", "van", "de", "der", "den", "del", "della", "di", "da", "dos", "du",
    "zu", "zur", "zum", "la", "le", "el", "al", "bin", "ibn", "ter", "ten",
}
# A leading INITIAL ("B. Winkler", "M. Schmidt-Rottluff"). Stripped before the
# sentence-punctuation test for the same reason an honorific's period is: it is
# an abbreviation mark, not a sentence boundary. Measured: without this, every
# initial+surname cell failed the shape gate outright, so the `Kürzel` column of
# a real workbook was invisible to the whole-cell override.
# Only the PERIOD is stripped here; the capitalization check below still runs on
# the original value, so a lowercase "b. winkler" is still rejected as a name.
_INITIAL_PREFIX = re.compile(r"^(?:[^\W\d_]\.[\s-]*)+")
# Stripped before the sentence-punctuation check only (kept in the value
# itself for the word/capitalization check below) -- an honorific's own
# period ("Dr. Klaus Müller") is not a sentence boundary. Kept in sync with
# core._HONORIFIC_PREFIX / engine._HONORIFICS.
_HONORIFIC_PREFIX = re.compile(r"^(?:Herrn?|Frau|Hr\.|Fr\.|Dr\.|Prof\.)\s+")
_SENTENCE_PUNCT = re.compile(r"[.!?]")
_MAX_NAME_CELL_LEN = 40


# A determiner/pronoun/verb/conjunction anywhere in an otherwise name-shaped
# value is strong evidence it's an ordinary phrase, not a name -- "Alle
# Zielwerte" ("All target values") and "Kein Ergebnis" ("No result") are both
# capitalized 2-word phrases that pass the shape check above, but neither is
# a name. Measured with the actual tagger: name particles ("von"/"de") tag as
# PROPN in a real name's context, so they're exempted rather than relied on to
# tag correctly under every phrasing.
_NON_NAME_POS = frozenset({"VERB", "AUX", "DET", "ADP", "CCONJ", "SCONJ", "PRON", "ADV", "NUM", "PUNCT", "SYM", "INTJ"})


def _looks_like_name(value: str, analyzer=None, lang: str | None = None) -> bool:
    """Gate for the whole-cell PERSON override below: a header substring match
    (_NAME_HEADER_TERMS is deliberately broad -- see its comment) is only
    evidence the COLUMN is about people, not that every cell in it is a bare
    name. Without this, a free-text prose column that happens to sit under a
    header containing e.g. "verantwortlich" gets every paragraph forcibly
    claimed as a person. A cell must be SHAPED like a name -- short, 1-4
    capitalized words, no sentence-ending punctuation -- AND (when an
    analyzer is available) contain no determiner/verb/conjunction/etc."""
    stripped = _INITIAL_PREFIX.sub("", _HONORIFIC_PREFIX.sub("", value))
    if len(value) >= _MAX_NAME_CELL_LEN or _SENTENCE_PUNCT.search(stripped):
        return False
    if "_" in value:
        return False  # snake_case field/status identifier ("Aktueller_Status"), not a name
    words = value.split()
    if not words or len(words) > 4:
        return False
    if not all(w.lower() in _NAME_PARTICLES or w[:1].isupper() for w in words):
        return False
    if analyzer is not None and lang is not None:
        try:
            doc = analyzer.nlp_engine.process_text(value, lang).tokens
        except Exception:  # noqa: BLE001 -- best-effort refinement; shape check alone already passed
            return True
        # Iterate spaCy's own tokens (not the whitespace-split `words` above --
        # a hyphenated name can tokenize differently) so a particle exemption
        # never depends on the two sequences lining up index-for-index.
        if doc is not None:
            for tok in doc:
                if tok.text.lower() in _NAME_PARTICLES:
                    continue
                if tok.pos_ in _NON_NAME_POS:
                    return False
    return True


# --- column-level name inference ---------------------------------------------
# The fourth corroboration source. The whole-cell override above needs the HEADER
# to say "people"; this one works out that a column holds people from its
# CONTENT, which is the only thing available when the header lies ("Status"),
# says nothing ("Feld_7"), or is missing entirely.
#
# Why it is worth the pass: measured on the hardened harness, a column of bare
# surnames under a misleading header scored 10%. NER catches the foreign and rare
# names in such a column (they score ~100% bare) and misses the everyday-word
# German ones (~25%) -- so the caught minority is evidence about the column, and
# that evidence rescues the missed majority. One column-level decision beats
# twenty independent cell-level coin flips.
#
# Why it is safe to treat as CORROBORATION rather than a bare guess: the trigger
# is an aggregate of several independent hits plus three shape gates, not one
# model opinion. The guards, each closing a measured failure mode:
_INFER_MIN_VALUES = 4          # fewer rows than this is not a distribution
_INFER_MIN_SHAPE_RATIO = 0.8   # the column must be name-SHAPED nearly throughout
_INFER_MIN_DISTINCT_RATIO = 0.5  # people columns list people; an ENUM repeats itself
_INFER_MIN_HITS = 2            # one stray PERSON hit is not a pattern
_INFER_MIN_HIT_RATIO = 0.2     # ...and it must be a real share of the column
_INFER_ROW_LIMIT = 500         # fixed cap, so the decision is identical every run


_VALIDATION_RANGE = re.compile(
    r"(?:'(?P<q>[^']+)'|(?P<b>[A-Za-z0-9_]+))?!?\$?(?P<c1>[A-Z]{1,3})\$?\d+(?::\$?(?P<c2>[A-Z]{1,3})\$?\d+)?"
)


def _validation_source_columns(wb) -> set[tuple[str, str]]:
    """{(sheet, column letter)} for columns that are the SOURCE LIST of a
    dropdown somewhere in the workbook.

    Such a column is a controlled vocabulary by the author's own declaration --
    "Idee / Validierung / Konzeption / Rollout" -- and never a roster of people,
    however name-shaped and however distinct its entries look. Measured: without
    this the inference below read the fixture's `DB_Setup` sheet (which exists
    only to back three real data validations) as four columns of people.

    Deliberately gates ONLY the inference, never the header override: a dropdown
    of EMPLOYEE names under a "Bearbeiter" header is a perfectly normal shape,
    and suppressing an explicitly declared people-column would be a leak."""
    out: set[tuple[str, str]] = set()
    for ws in wb.worksheets:
        try:
            validations = list(ws.data_validations.dataValidation)
        except Exception:  # noqa: BLE001 -- absent/oddly-shaped validations are not an error
            continue
        for dv in validations:
            formula = getattr(dv, "formula1", None)
            if not isinstance(formula, str) or not formula.strip():
                continue
            m = _VALIDATION_RANGE.search(formula.lstrip("="))
            if not m:
                continue
            sheet = m.group("q") or m.group("b") or ws.title
            cols = [c for c in (m.group("c1"), m.group("c2")) if c]
            if not cols:
                continue
            first, last = cols[0], cols[-1]
            for idx in range(column_index_from_string(first), column_index_from_string(last) + 1):
                out.add((sheet, get_column_letter(idx)))
    return out


def _inferred_name_columns(wb, analyzer, config, sheet_langs: dict[str, str]) -> set[tuple[str, str]]:
    """{(sheet title, column letter)} for columns whose CONTENT says they hold
    people, regardless of what their header says.

    Pure function of the workbook + config, so scan() and apply() derive the
    identical set and scan/apply parity holds by construction -- the same
    property `topical_gazetteer` and `_sheet_languages` rely on."""
    if not config.get("infer_name_columns", True):
        return set()
    doc_lang = (config.get("languages") or list(DEFAULT_LANGUAGES))[0]
    header_re = _name_header_re(tuple(config.get("name_column_headers", ())))
    vocabulary_cols = _validation_source_columns(wb)
    # A sheet that backs ANY dropdown is a lookup/setup sheet, and its OTHER
    # columns are vocabulary too. This is not over-reach, it is the shape such a
    # sheet has: measured on the fixture, only `DB_Setup!A` and `!B` are
    # reachable as declared sources (openpyxl drops the x14 extension that
    # carries the rest), yet `!C` -- "Idee / Validierung / Konzeption / Rollout"
    # -- is just as certainly a vocabulary and was inferred as a people column.
    # Restricted to the inference for the same reason as above: an explicit
    # people-header on such a sheet is still honoured.
    vocabulary_sheets = {sheet for sheet, _col in vocabulary_cols}
    out: set[tuple[str, str]] = set()
    for ws in wb.worksheets:
        lang = sheet_langs.get(ws.title, doc_lang)
        headers = _column_headers(ws)
        by_col: dict[int, list[str]] = {}
        for row in ws.iter_rows(min_row=2, max_row=1 + _INFER_ROW_LIMIT):
            for cell in row:
                if cell.value in (None, ""):
                    continue
                v = str(cell.value).strip()
                if v:
                    by_col.setdefault(cell.column, []).append(v)
        for col, values in by_col.items():
            header = headers.get(col, "")
            # A header that already says "people" is handled by the whole-cell
            # override; a header that declares a topical CATEGORY (Team, Abteilung)
            # is that category's column and must not be re-read as people.
            if header_re.search(header) or _category_for_header(header, config):
                continue
            if (ws.title, get_column_letter(col)) in vocabulary_cols or ws.title in vocabulary_sheets:
                continue  # the author declared this a dropdown vocabulary
            if len(values) < _INFER_MIN_VALUES:
                continue
            distinct = sorted(set(values))
            if len(distinct) / len(values) < _INFER_MIN_DISTINCT_RATIO:
                continue  # a controlled vocabulary, not a roster of people
            shaped = [v for v in distinct if _looks_like_name(v, analyzer, lang)]
            if len(shaped) / len(distinct) < _INFER_MIN_SHAPE_RATIO:
                continue
            hits = sum(1 for v in shaped if _is_person_value(v, analyzer, lang))
            if hits >= _INFER_MIN_HITS and hits / len(shaped) >= _INFER_MIN_HIT_RATIO:
                out.add((ws.title, get_column_letter(col)))
    return out


def _is_person_value(value: str, analyzer, lang: str) -> bool:
    """Independent evidence that ONE cell value names a person: the curated
    given-name gazetteer, or the model typing it PERSON on its own. Only used to
    count evidence ACROSS a column -- never on its own to claim a single cell."""
    if is_given_name(value):
        return True
    try:
        return any(r.entity_type == "PERSON" for r in analyzer.analyze(text=value, language=lang))
    except Exception:  # noqa: BLE001 -- inference is best-effort; a failure just means no vote
        return False


def _column_headers(ws) -> dict[int, str]:
    headers = {}
    first_row = next(ws.iter_rows(min_row=1, max_row=1), [])
    for cell in first_row:
        if isinstance(cell.value, str) and cell.value.strip():
            headers[cell.column] = cell.value.strip()
    return headers


def _cell_scan_text(cell) -> str | None:
    """The text to scan for a cell, or None to skip. String cells pass through;
    NUMBERS are coerced to a plain string so account / tax / customer / phone
    numbers STORED AS NUMBERS (very common in bank spreadsheets) are not
    invisible to detection -- previously only string cells were scanned, so a
    numeric account number sailed through into a "verified" file. Short numbers
    (< 5 digits: counts, small amounts) and non-integer decimals (monetary
    amounts) are skipped as not-identifiers; dates/booleans/formulas/errors are
    left to structure."""
    v = cell.value
    if v is None:
        return None
    if cell.data_type == "s" and isinstance(v, str):
        return v if v.strip() else None
    if cell.data_type == "n" and isinstance(v, (int, float)) and not isinstance(v, bool):
        if isinstance(v, float):
            if not v.is_integer():
                return None
            v = int(v)
        s = str(v)
        return s if len(s) >= 5 else None
    return None


# A formula's double-quoted STRING LITERALS. `=("Kunde "&"Mueller")` and
# `=CONCATENATE("Hans Mueller";A2)` put real PII in a cell whose stored value is a
# formula, which _cell_scan_text deliberately skips -- so it was never scanned,
# never removed, and (being undecided) never re-checked by the verify either.
# Only the literals are scanned and spliced: a token anywhere else in the
# expression would corrupt the formula, and the structure (function names, cell
# refs, operators) is not PII. `""` is Excel's escaped quote and stays inside the
# literal.
_FORMULA_LITERAL_RE = re.compile(r'"((?:[^"]|"")*)"')


def _formula_literals(formula: str) -> list[tuple[int, int, str]]:
    """(start, end, text) of each string literal in a formula, offsets into `formula`."""
    return [(m.start(1), m.end(1), m.group(1)) for m in _FORMULA_LITERAL_RE.finditer(formula)]


def _cell_formula(cell) -> str | None:
    return cell.value if cell.data_type == "f" and isinstance(cell.value, str) else None


def _iter_cell_units(wb):
    """Yields (id, text, header) -- header is the column-1-row text for that
    column, given as context so recognizers relying on nearby German context
    words (Kontonummer, Depotnummer, ...) actually have something to match,
    since a bare cell value alone carries no context.

    Row 1 itself is used ONLY as the header/schema label (via _column_headers)
    -- never scanned as its own data unit. A schema label ("NewValue",
    "Project_ID") is a structural name, not user data, and its usual
    CamelCase/underscore-joined shape reads as a proper noun to NER (measured:
    both get tagged PROPN, same as a real name), producing a finding that is
    really just the column's own name. Row-1 comments are still scanned --
    an actual annotation, unlike the label itself."""
    for ws in wb.worksheets:
        headers = _column_headers(ws)
        for row in ws.iter_rows():
            for cell in row:
                header = headers.get(cell.column) if cell.row != 1 else None
                if cell.row != 1:
                    text = _cell_scan_text(cell)
                    if text is not None:
                        yield f"cell|{ws.title}|{cell.coordinate}", text, header
                # Formula literals are scanned in EVERY row, row 1 included. The
                # row-1 exclusion above is about the schema LABEL -- a structural
                # name, not user data. A string literal embedded in a formula is
                # real content wherever it sits ( ="Kundennr "&A2 in a header row
                # is still a value), and it is invisible to the cell-text walk.
                formula = _cell_formula(cell)
                if formula is not None:
                    for k, (_s, _e, literal) in enumerate(_formula_literals(formula)):
                        if literal.strip():
                            yield f"formula|{ws.title}|{cell.coordinate}|{k}", literal, header
                if cell.comment is not None and cell.comment.text.strip():
                    yield f"comment|{ws.title}|{cell.coordinate}", cell.comment.text, header


def _iter_defined_name_units(wb):
    for name, defn in wb.defined_names.items():
        if isinstance(defn.value, str) and defn.value.strip():
            yield f"defined_name|{name}", defn.value


def _iter_sheet_name_units(wb):
    """The worksheet TITLES. Their authoritative copy is xl/workbook.xml
    <sheet name="...">, which no cell walk reaches -- and for this deployment a
    workbook with ONE SHEET PER CLIENT is a completely normal shape, so a sheet
    name is a first-class leak, not decoration. (docProps/app.xml TitlesOfParts is
    only a CACHE of these; blanking it removed the evidence, not the leak.)"""
    for ws in wb.worksheets:
        if ws.title.strip():
            yield f"sheetname|{ws.title}", ws.title


# The recognizer name presidio reports for a raw spaCy NER span -- i.e. a guess
# with nothing but the model behind it.
_BARE_NER_SOURCE = "SpacyRecognizer"


def _corroborated_for_sheet_title(f: Finding) -> bool:
    """Whether a finding is strong enough to justify RENAMING a worksheet.

    Renaming a sheet is structural, not textual: it rewrites xl/workbook.xml and
    every formula, defined name and chart reference that points at that sheet. It
    is a far heavier act than replacing a value inside one cell -- and it is driven
    by whatever detection makes of a string that is usually three to eight
    characters of structural vocabulary ("Tab", "Daten", "Q3", "Blatt2").

    Measured: spaCy claims a bare "Tab" as PERSON at its flat 0.85. PERSON is
    deliberately exempt from the corroboration-only rule AND from the
    lowercase/stopword precision filters -- both exemptions are correct, because a
    real lowercase surname has to stay reachable inside a CELL -- so nothing else
    stands between one weak three-character guess and a renamed tab with every
    reference to it rewritten.

    So a sheet title alone requires more than a bare spaCy guess. A pattern/anchor
    hit, a passed checksum, a deny-list term, a name-column override, or a value
    propagated from a confirmed detection elsewhere in the document all qualify.
    Non-NER entity types (IBAN, Steuer-ID, ...) are pattern-backed by definition.

    This does NOT weaken the leak this feature exists to close. A sheet actually
    called "Kunde Hans Mueller" still renames: "Hans Mueller" is corroborated by
    the name-column/propagation machinery the moment it occurs anywhere else in the
    workbook, and a title whose value the reviewer removed still redacts via the
    ordinary `redact()` trigger in _sheet_renames. What stops is renaming on the
    strength of nothing at all."""
    if f.entity_type not in _NER_ENTITIES:
        return True
    return f.validated is True or f.source != _BARE_NER_SOURCE


def extract_text_units(path: Path) -> list[TextUnit]:
    wb = openpyxl.load_workbook(path, data_only=False)
    units = [TextUnit(id=key, text=text) for key, text, _header in _iter_cell_units(wb)]
    units.extend(TextUnit(id=key, text=text) for key, text in _iter_defined_name_units(wb))
    units.extend(TextUnit(id=key, text=text) for key, text in _iter_sheet_name_units(wb))
    # Surfaces openpyxl's cell model never exposes: external hyperlink TARGETS
    # (the URL behind a cell link) and chart part text (see run_replace).
    units.extend(aux_text_units(path))
    return units


# --- column-level policy (redact/pseudonymize a whole column) -----------------
# A column policy blacks out EVERY non-empty cell in a column regardless of what
# detection found -- the only way to redact a column whose sensitivity is topical
# (a confidential project description) rather than an identifiable entity. Actions
# supported at the column level: "pseudonymize" (consistent per-column token) and
# "anonymize" (one-way). "skip" is intentionally NOT here: the value-keyed decision
# model can't express "keep this value here but remove it elsewhere", so a skipped
# column that shares a value with a removed one would trip the fail-loud verify.
_COLUMN_BLACKOUT_ACTIONS = ("pseudonymize", "anonymize", "summarize")


# German umlauts MUST be transliterated, not preserved: actions.TOKEN_RE is
# `[A-Z0-9_]` and can never match one, so a pseudonymized 'Prüfung' column wrote
# [PRÜFUNG_1] tokens that reidentify_text could NEVER reverse -- permanently
# unreversible, and silently so. The German transliteration keeps the label
# readable (PRUEFUNG, not PRFUNG) while staying inside the token alphabet.
_UMLAUT_TRANSLITERATION = str.maketrans(
    {"Ä": "AE", "Ö": "OE", "Ü": "UE", "ä": "AE", "ö": "OE", "ü": "UE", "ß": "SS", "ẞ": "SS"}
)


def _column_entity_type(header: str, col_letter: str) -> str:
    """Per-column entity type for pseudonym tokens, derived from the header so a
    pseudonymized 'Projekt' column renders readable, re-identifiable [PROJEKT_n]
    tokens. Falls back to the column letter when there is no header."""
    ascii_header = (header or "").strip().translate(_UMLAUT_TRANSLITERATION)
    base = re.sub(r"[^0-9A-Za-z]+", "_", ascii_header).strip("_").upper()
    return base or f"COLUMN_{col_letter}"


def _coord_column(coord: str) -> str | None:
    m = re.match(r"([A-Z]+)", coord)
    return m.group(1) if m else None


def cell_summary(findings: list) -> list[CellInfo]:
    """Every spreadsheet cell that carries a finding, so the reviewer can set a
    per-cell policy (the exception layer). Derived from the findings' unit_ids
    (`cell|Sheet|A5`) -- no workbook re-read -- with a short content preview and
    the detected entity types. Ordered by sheet then coordinate."""
    by_cell: dict[tuple[str, str], dict] = {}
    for f in findings:
        parts = f.unit_id.split("|")
        if len(parts) == 3 and parts[0] in ("cell", "comment"):
            sheet, coord = parts[1], parts[2]
            info = by_cell.setdefault((sheet, coord), {"types": set(), "sample": (f.context or f.value)})
            info["types"].add(f.entity_type)
    out = [
        CellInfo(sheet=s, coord=c, header="", sample=d["sample"], entity_types=tuple(sorted(d["types"])))
        for (s, c), d in by_cell.items()
    ]
    out.sort(key=lambda ci: (ci.sheet, column_index_from_string(_coord_column(ci.coord) or "A"), ci.coord))
    return out


def column_summary(path: Path, findings: list, config: dict | None = None) -> list[ColumnInfo]:
    """Describe each spreadsheet column (sheet, letter, header, a sample value,
    and how many findings landed in it) so the reviewer can set a whole-column
    policy. Only columns that carry a header OR at least one finding are listed --
    empty structural columns are noise."""
    header_re = _name_header_re(tuple((config or {}).get("name_column_headers", ())))
    counts: dict[tuple[str, str], int] = {}
    for f in findings:
        parts = f.unit_id.split("|")
        if len(parts) >= 3 and parts[0] in ("cell", "comment", "formula"):
            col = _coord_column(parts[2])
            if col:
                counts[(parts[1], col)] = counts.get((parts[1], col), 0) + 1

    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    out: list[ColumnInfo] = []
    try:
        for ws in wb.worksheets:
            headers: dict[str, str] = {}
            for cell in next(ws.iter_rows(min_row=1, max_row=1), []):
                if isinstance(cell.value, str) and cell.value.strip():
                    headers[get_column_letter(cell.column)] = cell.value.strip()
            wanted = set(headers) | {col for (sheet, col) in counts if sheet == ws.title}
            samples: dict[str, str] = {}
            for i, row in enumerate(ws.iter_rows(min_row=2)):
                if i >= 200 or len(samples) >= len(wanted):  # sample from the first rows only
                    break
                for cell in row:
                    if cell.value in (None, ""):
                        continue  # read_only fills row gaps with EMPTY_CELL, which has no .column
                    col = get_column_letter(cell.column)
                    if col in wanted and col not in samples:
                        samples[col] = str(cell.value)
            for col in sorted(wanted, key=column_index_from_string):
                header = headers.get(col, "")
                out.append(
                    ColumnInfo(
                        sheet=ws.title,
                        column=col,
                        header=header,
                        sample=samples.get(col, ""),
                        pii_count=counts.get((ws.title, col), 0),
                        name_override=bool(header_re.search(header)),
                    )
                )
    finally:
        wb.close()
    return out


def _combined_cell_text(text: str, header: str | None) -> str:
    prefix = f"{header}: " if header else ""
    return prefix + text


def _ml_scan_texts(wb) -> list[str]:
    """Every distinct text the ML pass will be asked about, in the exact form it will
    see it: the header+value combination, neutralized, i.e. what detect_unit hands to
    analyzer.analyze(). Derived from the workbook alone, so scan and apply produce the
    SAME list -- which is what makes priming parity-safe rather than a parity risk."""
    combined = {
        neutralize_structural_noise(_combined_cell_text(text, header))
        for _key, text, header in _iter_cell_units(wb)
    }
    combined |= {neutralize_structural_noise(text) for _key, text in _iter_defined_name_units(wb)}
    return sorted(combined)


def _precompute_cell_artifacts(wb, analyzer, config) -> dict[tuple[str | None, str], object]:
    """One spaCy pipe() batch over every DISTINCT (header, cell-text) combo in
    the workbook, instead of one analyze() call per cell -- measured ~5x faster
    for the many short, highly repetitive values a spreadsheet holds. Returns
    nlp_artifacts keyed by (header, text) so scan()/apply() can feed them
    straight into their existing per-cell cache. Empty (falls back to the
    per-call path) under a multi-language config, since a cached artifact is
    tied to one language."""
    languages = config.get("languages") or list(DEFAULT_LANGUAGES)
    if len(languages) != 1:
        return {}
    combined_by_key: dict[tuple[str | None, str], str] = {}
    for _, text, header in _iter_cell_units(wb):
        combined_by_key.setdefault((header, text), _combined_cell_text(text, header))
    for _, text in _iter_defined_name_units(wb):
        combined_by_key.setdefault((None, text), text)
    # Batch on the SAME cleaned text detect_unit will request via its own
    # neutralize_structural_noise call, so these precomputed artifacts (built
    # from a bullet/heading-fusion-neutralized copy) match what detect_unit
    # actually analyzes -- batching on the raw text here would precompute
    # tokenization for a string detect_unit never uses, silently discarding
    # the whole point of the cleanup.
    artifacts_by_clean = precompute_nlp_artifacts(
        analyzer, (neutralize_structural_noise(c) for c in combined_by_key.values()), languages[0]
    )
    return {key: artifacts_by_clean.get(neutralize_structural_noise(combined)) for key, combined in combined_by_key.items()}


def _analyze_cell_text(
    text: str,
    header: str | None,
    analyzer,
    config,
    unit_id: str = "tmp",
    nlp_artifacts=None,
    people_column: bool = False,
) -> list:
    combined = _combined_cell_text(text, header)
    unit = TextUnit(id=unit_id, text=combined)
    findings = detect_unit(analyzer, unit, config, nlp_artifacts=nlp_artifacts)
    offset = len(combined) - len(text)
    result = []
    for f in findings:
        if f.end <= offset:
            continue  # entirely inside the header context -- not the cell value
        if f.start < offset:
            # Span STRADDLES the header/value boundary: clip to the value side and
            # re-slice its value, rather than dropping it wholesale (which leaked the
            # in-value portion -- e.g. a deny term that included the header text).
            f.start = offset
            f.value = combined[f.start : f.end]
        f.start -= offset
        f.end -= offset
        result.append(f)

    # The column header declares this cell is a person. Trust it over the model --
    # including over a weaker guess that already claimed the whole cell (below).
    value = text.strip()
    # `value` is sliced from the REAL text (it has to be -- it becomes the
    # Finding's value and its span), but the "is there anything here at all"
    # question must be asked of the neutralized copy: a cell holding only Excel
    # `_xHHHH_` escapes is empty, yet reads as word characters. Same-length
    # neutralization means this probe never shifts an offset.
    # Measured: 7 empty cells were flagged DESCRIPTION at the auto-accept tier.
    probe = neutralize_structural_noise(value).strip()
    header_re = _name_header_re(tuple(config.get("name_column_headers", ())))
    languages = config.get("languages") or list(DEFAULT_LANGUAGES)
    lang = languages[0] if len(languages) == 1 else None
    # Either the header declares people, or the COLUMN'S CONTENT does
    # (_inferred_name_columns). Both are column-level evidence and both are
    # stronger than a per-cell model guess, so they share the same downstream
    # handling; only the source tag differs, so an audit can tell which fired.
    header_declares = bool(header_re.search(header or ""))
    column_says_people = header_declares or people_column
    override_source = "whole_cell_override" if header_declares else "inferred_name_column"
    header_says_people = bool(
        column_says_people
        and probe
        and not _NOT_A_NAME.match(probe)
        and _looks_like_name(value, analyzer, lang)
    )
    whole_cell = [f for f in result if f.start == 0 and f.end >= len(value)]
    if header_says_people:
        # A WEAKER guess must not silently veto the header. spaCy types some real
        # names NER_MISC rather than PERSON ("Constanza Hiemenz", measured on the
        # reported workbook): that MISC hit covers the whole cell, which suppressed
        # the override, and MISC -- being a bare guess -- was then dropped outright
        # by corroboration_only. Net effect, a name sitting in a column literally
        # headed "Owner" left the tool in the clear.
        #
        # Retyped IN PLACE rather than added alongside: a second finding on the same
        # span would lose the overlap contest to MISC's higher flat 0.85 and change
        # nothing. The column header is the stronger evidence about the TYPE, and
        # re-sourcing it also makes the value corroborated, which is what stops
        # corroboration_only from discarding it.
        for f in whole_cell:
            if f.entity_type in _CORROBORATION_ONLY_ENTITIES and f.source in _GATED_NER_SOURCES:
                f.entity_type = "PERSON"
                f.source = override_source
                f.score = max(f.score, _NAME_COLUMN_SCORE)
                f.context = f"{header}: {value}"
        if not whole_cell:
            start = text.index(value)
            result.append(
                Finding(
                    entity_type="PERSON",
                    value=value,
                    score=_NAME_COLUMN_SCORE,
                    context=f"{header}: {value}",
                    unit_id=unit_id,
                    start=start,
                    end=start + len(value),
                    source=override_source,
                )
            )

    # Topical header override: a column whose header maps to a category (Tool,
    # Abteilung, Lizenzgeber, Projekt, ...) makes the WHOLE cell that category --
    # the document's own schema is authoritative. Covers name columns (the cell
    # IS the tool/division name) and description columns (the whole PROJECT
    # description is claimed for redact/summarize). Source-tagged so it bypasses
    # the NER noise/corroboration filters (those gate only NER entity types).
    category = _category_for_header(header, config)
    if (
        category
        and probe
        and not _NOT_A_NAME.match(probe)
        and not _is_placeholder_token(probe)
        and not any(f.start == 0 and f.end >= len(value) for f in result)
    ):
        start = text.index(value)
        result.append(
            Finding(
                entity_type=category,
                value=value,
                score=_TOPICAL_SCORE,
                context=f"{header}: {value}",
                unit_id=unit_id,
                start=start,
                end=start + len(value),
                source="topical_header",
            )
        )

    # The whole-cell override can PARTIALLY overlap a finding NER did make (just the
    # surname, or a KONTO number in the same cell). Appending it raw left overlapping
    # spans, which the cell splicer assumes never happens -> garbled tokens. Re-resolve
    # the combined set so the no-overlap invariant holds (the override merges to cover
    # the cell rather than corrupting it).
    return _resolve_overlaps(result, text)


def _sheet_languages(wb, config) -> dict[str, str]:
    """{sheet title: language} -- the language each SHEET is scanned in.

    Language was decided once per DOCUMENT, which is wrong for the shape a bank
    workbook actually has: one file with German sheets and an English client
    register in it. Measured on the audit workbook (2026-07-26): the whole file
    routed to German, so every `languages: [en]` recognizer stayed unregistered and
    ten GDPR Art. 9 values on the English sheet -- health conditions, union
    memberships, ethnic origins -- were never detected, while the German Art. 9
    word lists ran on the English prose and claimed "The Great Depression started
    in 1929" as health data. One decision, both failure directions.

    A sheet is a coherent language unit in a way a workbook is not, so the decision
    moves there. Only a CONFIDENT detection overrides the document language; a
    sheet of bare numbers or names has no language signal and must not be routed on
    a coin-flip. Pure function of the workbook's text, so scan and apply derive the
    identical map and parity holds by construction."""
    doc_lang = (config.get("languages") or list(DEFAULT_LANGUAGES))[0]
    supported = set(config.get("languages") or ()) | set(DEFAULT_LANGUAGES)
    by_sheet: dict[str, list[str]] = {}
    for key, text, _header in _iter_cell_units(wb):
        sheet = key.split("|", 2)[1]
        bucket = by_sheet.setdefault(sheet, [])
        if len(bucket) < 400:  # bounded: a language read needs a sample, not the sheet
            bucket.append(text)
    out: dict[str, str] = {}
    for ws in wb.worksheets:
        # HEADERS are part of the sample, and matter more than anything else here.
        # _iter_cell_units deliberately skips row 1 (a schema label is not user
        # data), but on a table the header row is the strongest language signal
        # there is -- the body is mostly names, numbers and IBANs, which read as no
        # language at all. Measured: without the headers the English client
        # register was not confidently English and fell back to the document
        # language, which is the whole failure this function exists to fix.
        sample = " ".join([*_column_headers(ws).values(), *by_sheet.get(ws.title, [])])
        lang, _confident = language.detect_dominant(sample) if sample.strip() else (doc_lang, False)
        # The confidence flag is deliberately NOT required here, unlike the
        # document-level routing it complements. A value TABLE -- surnames, IBANs,
        # nationalities -- has almost no function words, so it is essentially never
        # "confident" while still being clearly one language or the other: measured,
        # the English client register reads ('en', False) and the German client
        # table reads ('de', False), both correct in direction. Demanding confidence
        # threw the direction away and sent every table to the document language,
        # which is the failure this exists to fix. The document language remains the
        # fallback for a sheet with no text at all.
        out[ws.title] = lang if lang in supported else doc_lang
    return out


# A cell needs at least this many characters before its own text is a better
# language signal than the sheet it sits on. Below it, a value ("Saldo", "Klaus
# Mueller", "EUR 4,2 Mio.") has no language at all and the sheet must decide.
_TEXT_LANG_MIN_CHARS = 40


@functools.lru_cache(maxsize=8192)
def _text_language(text: str, fallback: str, supported: tuple[str, ...]) -> str:
    """The language of ONE cell, when the cell is long enough to have one.

    Sheet-level routing is right for tables but wrong for the sheet that is a pile
    of PROSE in both languages -- a strategy or minutes sheet where German and
    English paragraphs sit in the same column. Measured: routing such a sheet to
    English (correctly, most of it is) then ran the ENGLISH spaCy model over its
    German sentences and produced exactly the noise per-document narrowing exists
    to prevent -- "Konditionen", "Saldo" and "Zum Stichtag" all claimed as PERSON.

    A long free-text cell carries its own signal, so it decides for itself. This
    RAISES recall rather than trading it: each paragraph is scanned by the language
    whose recognizers actually apply to it. Pure function of the text, so scan and
    apply agree and parity holds; lru_cached because a spreadsheet repeats text.

    Confidence IS required here, unlike the sheet-level routing which deliberately
    drops it. The two cases differ: a sheet aggregates hundreds of cells and is
    never "confident" while still being reliably directional, whereas ONE sentence
    that the detector is unsure about is a coin flip. Measured at this length --
    every confident verdict was correct ("Im Anhang 2 sind die Konditionen
    beschrieben." -> de, "Race conditions in the settlement engine were fixed."
    -> en) while the unconfident ones included a plainly wrong one ("Credit Union:
    Nationwide Building Society" -> de). An unsure cell defers to its sheet.
    """
    if len(text) < _TEXT_LANG_MIN_CHARS:
        return fallback
    lang, confident = language.detect_dominant(text)
    return lang if (confident and lang in supported) else fallback


def _cfg_for_lang(config: dict, lang: str) -> dict:
    """The config to detect with, narrowed to one language."""
    if config.get("languages") == [lang]:
        return config
    return {**config, "languages": [lang]}


def scan(path: Path, analyzer, config) -> list:
    wb = openpyxl.load_workbook(path, data_only=False)
    findings = []
    sheet_langs = _sheet_languages(wb, config)
    doc_lang = (config.get("languages") or list(DEFAULT_LANGUAGES))[0]
    supported = tuple(sorted(set(config.get("languages") or ()) | set(DEFAULT_LANGUAGES)))
    # A "database" sheet repeats the same value thousands of times (a status, a
    # division, a city), and detection (one spaCy NER pass per cell) is the entire
    # cost. Memoize by (header, cell-text) for this scan: identical cells detect
    # once. Findings are re-stamped with each cell's unit_id (offsets/values are
    # relative to the cell text, so nothing else changes) so completeness-scan
    # coverage still maps to the right unit.
    # Keyed by LANGUAGE too: the same string on a German and an English sheet is
    # two different detection problems, and sharing one cache entry between them
    # would silently give whichever sheet ran second the other one's answer.
    # Keyed by the people-column flag as well: the same (header, text) pair can
    # legitimately appear in an inferred people column on one sheet and an
    # ordinary column on another, and sharing one cache entry would give
    # whichever ran second the other one's answer.
    cache: dict[tuple[str, str | None, str, bool], list] = {}
    artifacts_by_key = _precompute_cell_artifacts(wb, analyzer, config)
    # Batch the ML pass over the same text set BOTH passes derive (see _ml_scan_texts).
    prime_gliner(analyzer, _ml_scan_texts(wb))
    people_cols = _inferred_name_columns(wb, analyzer, config, sheet_langs)

    def detect(text, header, key, sheet=None, people_column=False):
        lang = _text_language(text, sheet_langs.get(sheet or "", doc_lang), supported)
        base = cache.get((lang, header, text, people_column))
        if base is None:
            # Precomputed artifacts are tied to the DOCUMENT language, so they are
            # only reusable where that is the language actually being used;
            # elsewhere detect_unit does its own NLP pass in the right language
            # rather than reuse a mismatched tokenization.
            arts = artifacts_by_key.get((header, text)) if lang == doc_lang else None
            base = _analyze_cell_text(
                text,
                header,
                analyzer,
                _cfg_for_lang(config, lang),
                nlp_artifacts=arts,
                people_column=people_column,
            )
            cache[(lang, header, text, people_column)] = base
        return [replace(f, unit_id=key) for f in base]

    for key, text, header in _iter_cell_units(wb):
        parts = key.split("|")
        sheet = parts[1]
        col = _coord_column(parts[2]) if len(parts) > 2 else None
        findings.extend(
            detect(text, header, key, sheet, people_column=(sheet, col) in people_cols)
        )
    for key, text in _iter_defined_name_units(wb):
        findings.extend(detect(text, None, key))
    for key, text in _iter_sheet_name_units(wb):
        # Filtered AFTER the shared memo cache, never inside it: the same string can
        # also be a cell value, where a bare NER guess is still a legitimate finding.
        findings.extend(f for f in detect(text, None, key) if _corroborated_for_sheet_title(f))
    # scan() builds its own unit stream rather than reusing extract_text_units, so
    # the auxiliary surfaces must be added here too or they would be reported to
    # the reviewer but never actually scanned.
    for unit in aux_text_units(path):
        findings.extend(detect_unit(analyzer, unit, config))
    return findings


def _apply_findings_to_text(
    text: str, header: str | None, analyzer, config, decisions: dict, mapping_store,
    nlp_artifacts=None, people_column: bool = False,
) -> str:
    findings = _analyze_cell_text(
        text, header, analyzer, config, nlp_artifacts=nlp_artifacts, people_column=people_column
    )
    if not findings:
        return text
    result = text
    for f in sorted(findings, key=lambda f: -f.start):
        action = decisions_lookup(decisions, f.entity_type, f.value)
        replacement = resolve_replacement(f.entity_type, f.value, action, mapping_store)
        if replacement is None:
            continue
        result = result[: f.start] + replacement + result[f.end :]
    return result


# --- sheet-name redaction -----------------------------------------------------
#
# Excel forbids [ ] : * ? / \ in a worksheet title and caps it at 31 characters,
# so a redacted title CANNOT carry a "[PERSON_1]" token -- there is no legal way to
# write one. A title that holds values the reviewer removed is therefore replaced
# wholesale by a neutral, unique placeholder. Deliberate trade-off: the TITLE is
# one-way even when the value in it was pseudonymized. Nothing becomes
# unrecoverable that was not already: the same person's name still appears as its
# pseudonym everywhere it occurs in cells, and the alternative -- leaving the tab
# reading "Kunde Hans Mueller" -- is the leak this exists to close.
_SHEET_TITLE_PLACEHOLDERS = ("Sheet", "Blatt", "Tabelle", "Anonym")
_MAX_SHEET_TITLE = 31


def _checked_removed_values(decisions: dict, config) -> set[str]:
    """The lower-cased values whose literal survival anywhere in the output is a
    leak, using the same rule as pipeline._literal_residual: everything the reviewer
    removed, minus values under 4 characters (too short to identify anyone and a
    magnet for substring false positives), plus every deny-list term regardless of
    length. Used to prove a REPLACEMENT title is itself clean."""
    removed = {
        value
        for (_entity, value), action in decisions.items()
        if action in _COLUMN_BLACKOUT_ACTIONS and len(value) >= 4
    }
    removed.update(t.strip().lower() for t in (config.get("deny_list") or []) if t.strip())
    return removed


def _sheet_renames(wb, redact, removed: set[str]) -> dict[str, str]:
    """{old title: new title} for every sheet whose NAME carries a removed value.

    The trigger is the ordinary redaction: if redacting the title changes it, the
    title holds something the reviewer removed. It has to be exactly that test and
    not a length-filtered one, because pipeline.verify_output re-scans the output
    with the recognizers and would re-detect a short value just the same.

    SECOND trigger, and the one that makes this agree with the verifier: a title
    that literally CONTAINS a checked removed value is renamed even when redacting
    it changes nothing. Those two tests can disagree, and did -- measured on the
    audit workbook (2026-07-26). A hidden sheet called "Archiv Thomas Weber" is
    claimed by spaCy as ONE PERSON span covering the whole title, so the only value
    ever offered for decision was "Archiv Thomas Weber"; the reviewer decided
    "Thomas Weber" (from the cells) instead, redact() therefore left the title
    alone, and the name shipped in xl/workbook.xml. verify_output caught it and
    blocked the save -- correct, but it meant a sheet named after a customer made
    the whole workbook impossible to process, with an error naming a value the
    reviewer had already dealt with. Keying the rename on exactly what the verifier
    checks removes the disagreement by construction.

    The replacement is a neutral placeholder PROVEN clean two ways -- no decided
    value redacts out of it, and no checked removed value is a substring of it. A
    workbook whose sheet is literally called "Sheet" would otherwise be renamed to
    "Sheet_1" and still contain the removed value, so such candidates are rejected
    and the next stem is tried. If none is safe we fail loud rather than ship it."""
    taken = set(wb.sheetnames)
    renames: dict[str, str] = {}
    for ws in wb.worksheets:
        title = ws.title
        title_low = title.lower()
        carries_removed = any(v in title_low for v in removed)
        if not title.strip() or (redact(title, None) == title and not carries_removed):
            continue
        candidate = None
        for stem in _SHEET_TITLE_PLACEHOLDERS:
            for n in range(1, len(wb.worksheets) + 50):
                trial = f"{stem}_{n}"
                if trial in taken or len(trial) > _MAX_SHEET_TITLE:
                    continue
                trial_low = trial.lower()
                if any(v in trial_low for v in removed) or redact(trial, None) != trial:
                    continue
                candidate = trial
                break
            if candidate:
                break
        if candidate is None:
            raise ProcessingError(
                f"Could not build a safe replacement name for sheet '{title}' -- every "
                "candidate still contained a value that had to be removed. No file was written."
            )
        taken.add(candidate)
        renames[title] = candidate
    return renames


def _rewrite_sheet_refs(expr: str, renames: dict[str, str]) -> str:
    """Repoints every sheet reference in a formula / defined-name expression at the
    renamed sheet. Both spellings Excel uses are handled: the quoted `'Kunde X'!A1`
    (internal apostrophes doubled) and the bare `Umsatz!A1`. The replacement is
    always emitted quoted, which is valid for any title."""
    for old, new in renames.items():
        quoted_old = "'" + old.replace("'", "''") + "'"
        expr = expr.replace(f"{quoted_old}!", f"'{new}'!")
        # Bare form only when nothing sheet-name-ish precedes it, so "BLA!A1" is
        # never mangled by a sheet called "A".
        expr = re.sub(rf"(?<![\w.']){re.escape(old)}!", f"'{new}'!", expr)
    return expr


def _apply_sheet_renames(wb, renames: dict[str, str]) -> None:
    if not renames:
        return
    for ws in wb.worksheets:
        new = renames.get(ws.title)
        if new is not None:
            ws.title = new
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                formula = _cell_formula(cell)
                if formula is not None:
                    rewritten = _rewrite_sheet_refs(formula, renames)
                    if rewritten != formula:
                        cell.value = rewritten
    for _name, defn in wb.defined_names.items():
        if isinstance(defn.value, str):
            rewritten = _rewrite_sheet_refs(defn.value, renames)
            if rewritten != defn.value:
                defn.value = rewritten


def _assert_no_stale_sheet_refs(wb, renames: dict[str, str]) -> None:
    """Fail loud if a renamed sheet's OLD title still appears in any formula or
    defined name. Two things go wrong at once if it does: the reference is now
    dangling (a #REF! in the colleague's copy), and the old title IS the PII we
    just removed -- so shipping the file would be the leak. Reference spellings
    this rewriter does not know (3D refs across a sheet RANGE) land here."""
    if not renames:
        return
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                formula = _cell_formula(cell)
                if formula and any(old in formula for old in renames):
                    raise ProcessingError(
                        f"Sheet '{ws.title}' cell {cell.coordinate} still references a sheet name "
                        "that had to be redacted, in a form this tool cannot rewrite safely. "
                        "No file was written."
                    )
    for name, defn in wb.defined_names.items():
        if isinstance(defn.value, str) and any(old in defn.value for old in renames):
            raise ProcessingError(
                f"Defined name '{name}' still references a sheet name that had to be redacted, "
                "in a form this tool cannot rewrite safely. No file was written."
            )


def apply(path: Path, out_path: Path, decisions: dict, analyzer, config, mapping_store) -> None:
    # keep_vba=False (the default) strips any macro project from the output,
    # which is intentional: anonymized copies are never macro-enabled.
    wb = openpyxl.load_workbook(path, data_only=False, keep_vba=False)
    # Whole-column blackout policies: {"Sheet!A": "pseudonymize"|"anonymize"|"summarize"}.
    column_policies = config.get("column_policies", {}) or {}
    # Per-CELL policies: {"Sheet!A5": mode} -- the finest-grained EXCEPTION layer,
    # wins over the column policy for that one cell. Blackout modes only (same
    # verify-safety constraint as column policy: a value-keyed "skip" can't be
    # expressed out-of-band without tripping the fail-loud residual check).
    cell_policies = config.get("cell_policies", {}) or {}
    blackout_cache: dict[tuple[str, str], str] = {}  # (key, value) -> token

    def blackout(col_key: str, header: str, col_letter: str, value: str, action: str) -> str:
        cached = blackout_cache.get((col_key, value))
        if cached is None:
            entity = _column_entity_type(header, col_letter)
            cached = resolve_replacement(entity, value, action, mapping_store) or value
            blackout_cache[(col_key, value)] = cached
        return cached

    # Memoize the redacted output by (header, text) for this apply -- a repeated
    # cell redacts to the same string. Safe: the pseudonym mapping is value-keyed,
    # so the same value already maps to the same token whether recomputed or cached
    # (the first call creates the mapping entry; the rest reuse the string).
    # Keyed by LANGUAGE too, for the same reason scan()'s cache is: the identical
    # string on a German and an English sheet is two different problems.
    redact_cache: dict[tuple[str, str | None, str, bool], str] = {}
    artifacts_by_key = _precompute_cell_artifacts(wb, analyzer, config)
    # Batch the ML pass over the same text set BOTH passes derive (see _ml_scan_texts).
    prime_gliner(analyzer, _ml_scan_texts(wb))
    sheet_langs = _sheet_languages(wb, config)
    doc_lang = (config.get("languages") or list(DEFAULT_LANGUAGES))[0]
    supported = tuple(sorted(set(config.get("languages") or ()) | set(DEFAULT_LANGUAGES)))
    # Derived here, BEFORE the sheet renames below, so the keys are the original
    # titles -- exactly what scan() saw. Deriving it after would key the set to
    # redacted titles and quietly disagree with scan on every renamed sheet,
    # which is a scan/apply parity break: findings the reviewer approved would
    # not be reproduced at apply and the fail-loud residual check would fire.
    people_cols = _inferred_name_columns(wb, analyzer, config, sheet_langs)

    def redact(
        text: str, header: str | None, sheet: str | None = None, people_column: bool = False
    ) -> str:
        lang = _text_language(text, sheet_langs.get(sheet or "", doc_lang), supported)
        out = redact_cache.get((lang, header, text, people_column))
        if out is None:
            arts = artifacts_by_key.get((header, text)) if lang == doc_lang else None
            out = _apply_findings_to_text(
                text, header, analyzer, _cfg_for_lang(config, lang),
                decisions, mapping_store, nlp_artifacts=arts, people_column=people_column,
            )
            redact_cache[(lang, header, text, people_column)] = out
        return out

    def redact_formula(
        formula: str, header: str | None, sheet: str | None = None, people_column: bool = False
    ) -> str:
        """Redacts only the quoted string literals, right-to-left so earlier
        offsets stay valid -- the expression around them is left byte-identical."""
        result = formula
        for start, end, literal in reversed(_formula_literals(formula)):
            if not literal.strip():
                continue
            new_literal = redact(literal, header, sheet, people_column)
            if new_literal != literal:
                result = result[:start] + new_literal + result[end:]
        return result

    # Sheet TITLES are redacted before anything else, so the reference rewrite runs
    # while formulas still hold the original names -- and every later pass then
    # sees the already-safe title.
    renames = _sheet_renames(wb, redact, _checked_removed_values(decisions, config))
    _apply_sheet_renames(wb, renames)
    # Column policies were keyed by the reviewer against the ORIGINAL sheet titles,
    # so a renamed sheet must still find its policy.
    original_title = {new: old for old, new in renames.items()}

    for ws in wb.worksheets:
        sheet_key = original_title.get(ws.title, ws.title)
        headers = _column_headers(ws)
        for row in ws.iter_rows():
            for cell in row:
                col_letter = get_column_letter(cell.column)
                header = headers.get(cell.column) if cell.row != 1 else None
                # Per-CELL policy first (finest granularity) -- wins over the
                # column policy and any per-value decision for this one cell.
                # Keyed on sheet_key -- the ORIGINAL title -- for the same reason
                # the column policy is: sheet renames are applied above, so by
                # this line ws.title is already the redacted name, and a decision
                # the reviewer recorded against the title they actually saw would
                # silently resolve to nothing on exactly the sheets that carry a
                # sensitive name.
                cell_policy = cell_policies.get(f"{sheet_key}!{cell.coordinate}")
                if (
                    cell_policy in _COLUMN_BLACKOUT_ACTIONS
                    and cell.row != 1
                    and cell.data_type in ("s", "n")
                    and cell.value not in (None, "")
                ):
                    cell.value = blackout(
                        f"{sheet_key}!{cell.coordinate}", headers.get(cell.column, ""), col_letter,
                        str(cell.value), cell_policy,
                    )
                    continue
                policy = column_policies.get(f"{sheet_key}!{col_letter}")
                # A column blackout wins over any per-value decision: EVERY non-empty
                # cell (header row excluded) is replaced, including cells detection
                # never flagged. Formula cells are left to the value path below,
                # which redacts the PII inside their string literals without
                # destroying the expression.
                if (
                    policy in _COLUMN_BLACKOUT_ACTIONS
                    and cell.row != 1
                    and cell.data_type in ("s", "n")
                    and cell.value not in (None, "")
                ):
                    cell.value = blackout(
                        f"{sheet_key}!{col_letter}", headers.get(cell.column, ""), col_letter, str(cell.value), policy
                    )
                    continue
                # Row 1 is the schema label, never scanned as its own data unit
                # (see _iter_cell_units) -- excluded here too so apply() redacts
                # exactly what scan() surfaced, never more (scan/apply parity).
                is_people_col = (sheet_key, col_letter) in people_cols
                text = _cell_scan_text(cell) if cell.row != 1 else None
                if text is not None:
                    new_value = redact(text, header, sheet_key, is_people_col)
                    if new_value != text:
                        # A redacted numeric cell must become a string cell so
                        # the token ("[KONTO_1]") can be stored at all.
                        cell.value = new_value
                formula = _cell_formula(cell)
                if formula is not None:
                    new_formula = redact_formula(formula, header, sheet_key, is_people_col)
                    if new_formula != formula:
                        cell.value = new_formula
                if cell.comment is not None and cell.comment.text.strip():
                    new_text = redact(cell.comment.text, header, sheet_key, is_people_col)
                    if new_text != cell.comment.text:
                        cell.comment.text = new_text
    for name, defn in wb.defined_names.items():
        if isinstance(defn.value, str) and defn.value.strip():
            new_value = redact(defn.value, None)
            if new_value != defn.value:
                defn.value = new_value
    # Checked LAST, once the string literals inside formulas have been redacted, so
    # a leftover really is an unrewritten reference rather than redacted prose.
    _assert_no_stale_sheet_refs(wb, renames)
    wb.save(out_path)
    apply_aux_parts(out_path, analyzer, config, decisions, mapping_store)
